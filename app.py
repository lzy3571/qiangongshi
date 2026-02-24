from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session as flask_session
from database import Base, Session, Mechanic, MonthlyRecord, RewardHistory, Issue, ClearedHoursRecord, User, Attachment6Data, Attachment1Data, Workshop, Team, RouteHoursData, OperationLog, engine
from sqlalchemy import func, desc, asc, inspect, text, or_, and_
from sqlalchemy.orm import joinedload
import pandas as pd
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import shutil
import json

import re

# Login Manager
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user

app = Flask(__name__)
app.secret_key = 'supersecretkey'
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = None

def ensure_issue_status_column():
    try:
        inspector = inspect(engine)
        cols = [c['name'] for c in inspector.get_columns('issues')]
        if 'status' not in cols:
            with engine.connect() as conn:
                conn.execute(text("ALTER TABLE issues ADD COLUMN status VARCHAR DEFAULT '未结算'"))
                conn.commit()
    except Exception:
        pass

ensure_issue_status_column()

def ensure_operation_logs_table():
    try:
        inspector = inspect(engine)
        if 'operation_logs' not in inspector.get_table_names():
            with engine.connect() as conn:
                conn.execute(text("CREATE TABLE IF NOT EXISTS operation_logs (id INTEGER PRIMARY KEY AUTOINCREMENT, user VARCHAR, action VARCHAR, detail VARCHAR, created_at DATETIME)"))
                conn.commit()
    except Exception:
        pass

ensure_operation_logs_table()

Base.metadata.create_all(engine)

def ensure_att6_consecutive_info():
    try:
        session = Session()

        def normalize_month(val):
            if not val:
                return None
            s = str(val).strip()
            if not s:
                return None
            s = s.replace('年', '-').replace('月', '').replace('.', '-').replace('/', '-')
            parts = s.split('-')
            if len(parts) >= 2:
                y = parts[0]
                m = parts[1]
                if len(m) == 1:
                    m = f"0{m}"
                return f"{y}-{m}"
            if len(s) == 6 and s.isdigit():
                return f"{s[:4]}-{s[4:]}"
            return None

        def cycle_end_month(val):
            if not val:
                return None
            end_part = str(val).split('-')[-1].strip()
            return normalize_month(end_part)

        employee_ids = [r[0] for r in session.query(Attachment6Data.employee_id).distinct().all() if r and r[0]]
        if not employee_ids:
            session.close()
            return

        mech_rows = session.query(Mechanic.id, Mechanic.employee_id).filter(Mechanic.employee_id.in_(employee_ids)).all()
        employee_to_mid = {r[1]: r[0] for r in mech_rows if r and r[0] and r[1]}
        mids = list({mid for mid in employee_to_mid.values() if mid})
        if not mids:
            session.close()
            return

        hist_rows = session.query(RewardHistory.mechanic_id, RewardHistory.reward_date, RewardHistory.extra_reward).filter(
            RewardHistory.mechanic_id.in_(mids)
        ).order_by(RewardHistory.mechanic_id.asc(), RewardHistory.reward_date.asc()).all()

        per_mid = {}
        for mid, rdate, extra in hist_rows:
            mkey = normalize_month(rdate) or str(rdate or '').strip()
            if not mkey:
                continue
            per_mid.setdefault(mid, []).append((mkey, float(extra or 0.0)))

        prev_extra_map = {}
        for mid, arr in per_mid.items():
            seen = {}
            for mkey, extra in arr:
                seen[mkey] = extra
            keys_sorted = sorted(seen.keys())
            prev = 0.0
            for k in keys_sorted:
                prev_extra_map[(mid, k)] = prev
                prev = float(seen[k] or 0.0)

        att6_all = session.query(Attachment6Data).filter(Attachment6Data.employee_id.in_(employee_ids)).all()
        for r in att6_all:
            mid = employee_to_mid.get(r.employee_id)
            mkey = normalize_month(r.reward_date) or cycle_end_month(r.reward_cycle) or str(r.reward_date or '').strip()
            prev_extra = prev_extra_map.get((mid, mkey), 0.0)
            r.consecutive_info = f"{float(prev_extra or 0.0):.2f}"

        session.commit()
        session.close()
    except Exception:
        try:
            session.rollback()
            session.close()
        except Exception:
            pass

ensure_att6_consecutive_info()

def write_operation_log(action, detail, session=None, user=None):
    try:
        own_session = session is None
        log_session = session or Session()
        actor = user if user is not None else (current_user.username if current_user.is_authenticated else 'system')
        log = OperationLog(user=actor, action=action, detail=detail, created_at=datetime.now())
        log_session.add(log)
        if own_session:
            log_session.commit()
            log_session.close()
    except Exception:
        if session is None:
            try:
                log_session.rollback()
                log_session.close()
            except Exception:
                pass

def is_in_month_common(date_str, target_month):
    if not date_str:
        return False
    try:
        d_s = str(date_str)
        if d_s.startswith(target_month):
            return True
        target_slash = target_month.replace('-', '/')
        if d_s.startswith(target_slash):
            return True
        if d_s.replace('.', '').isdigit() and float(d_s) > 30000:
            dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(float(d_s)) - 2)
            return dt.strftime('%Y-%m') == target_month
        target_dot = target_month.replace('-', '.')
        if d_s.startswith(target_dot):
            return True
        return False
    except:
        return False

class UserObj(UserMixin):
    def __init__(self, id, username, role, workshop_id):
        self.id = id
        self.username = username
        self.role = role
        self.workshop_id = workshop_id

@login_manager.user_loader
def load_user(user_id):
    session = Session()
    user = session.query(User).get(int(user_id))
    session.close()
    if user:
        return UserObj(user.id, user.username, user.role, user.workshop_id)
    return None

ROLE_SECTION = 'section'
ROLE_WORKSHOP = 'workshop'

@app.context_processor
def inject_roles():
    return dict(ROLE_SECTION=ROLE_SECTION, ROLE_WORKSHOP=ROLE_WORKSHOP)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        session = Session()
        user = session.query(User).filter_by(username=username).first()
        session.close()
        
        if user and user.password == password: # In production use hashing!
            user_obj = UserObj(user.id, user.username, user.role, user.workshop_id)
            login_user(user_obj)
            return redirect(url_for('index'))
        else:
            flash('用户名或密码错误', 'error')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/ping')
def ping():
    return jsonify({"status": "ok", "message": "Server is running"})

@app.route('/users')
@login_required
def manage_users():
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('index'))
        
    session = Session()
    users = session.query(User).options(joinedload(User.workshop)).all()
    workshops = session.query(Workshop).all()
    session.close()
    return render_template('users.html', users=users, workshops=workshops)

@app.route('/users/add', methods=['POST'])
@login_required
def add_user():
    if current_user.role != ROLE_SECTION:
        return redirect(url_for('index'))
        
    username = request.form['username']
    password = request.form['password']
    role = request.form.get('role', ROLE_WORKSHOP)
    
    # New Fields
    employee_id = request.form.get('employee_id')
    name = request.form.get('name')
    contact = request.form.get('contact')
    workshop_id = request.form.get('workshop_id')
    
    session = Session()
    if session.query(User).filter_by(username=username).first():
        flash('用户名已存在', 'error')
    else:
        new_user = User(
            username=username, 
            password=password, 
            role=role,
            employee_id=employee_id,
            name=name,
            contact=contact,
            workshop_id=int(workshop_id) if workshop_id else None
        )
        session.add(new_user)
        session.commit()
        flash('用户添加成功', 'success')
    session.close()
    return redirect(url_for('manage_users'))

# --- Workshop & Team Management ---

@app.route('/data/workshops')
@login_required
def data_workshops():
    session = Session()
    workshops = session.query(Workshop).all()
    session.close()
    return render_template('data_workshops.html', workshops=workshops)

@app.route('/data/workshops/add', methods=['POST'])
@login_required
def add_workshop():
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('data_workshops'))
        
    name = request.form.get('name')
    if name:
        session = Session()
        try:
            session.add(Workshop(name=name))
            session.commit()
            flash('车间添加成功', 'success')
        except:
            session.rollback()
            flash('添加失败，可能名称重复', 'error')
        session.close()
    return redirect(url_for('data_workshops'))

@app.route('/data/workshops/delete/<int:id>')
@login_required
def delete_workshop(id):
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('data_workshops'))
        
    session = Session()
    w = session.query(Workshop).get(id)
    if w:
        session.delete(w)
        session.commit()
        flash('车间已删除', 'success')
    session.close()
    return redirect(url_for('data_workshops'))

@app.route('/data/teams')
@login_required
def data_teams():
    session = Session()
    teams = session.query(Team).options(joinedload(Team.workshop)).all()
    workshops = session.query(Workshop).all()
    session.close()
    return render_template('data_teams.html', teams=teams, workshops=workshops)

@app.route('/data/teams/add', methods=['POST'])
@login_required
def add_team():
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('data_teams'))
        
    name = request.form.get('name')
    workshop_id = request.form.get('workshop_id')
    
    if name:
        session = Session()
        try:
            session.add(Team(name=name, workshop_id=int(workshop_id) if workshop_id else None))
            session.commit()
            flash('班组添加成功', 'success')
        except Exception as e:
            session.rollback()
            flash(f'添加失败: {e}', 'error')
        session.close()
    return redirect(url_for('data_teams'))

@app.route('/data/teams/delete/<int:id>')
@login_required
def delete_team(id):
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('data_teams'))
        
    session = Session()
    t = session.query(Team).get(id)
    if t:
        session.delete(t)
        session.commit()
        flash('班组已删除', 'success')
    session.close()
    return redirect(url_for('data_teams'))

@app.route('/users/delete/<int:id>')
@login_required
def delete_user(id):
    if current_user.role != ROLE_SECTION:
        return redirect(url_for('index'))
        
    session = Session()
    user = session.query(User).get(id)
    if user:
        if user.username == 'admin':
            flash('无法删除管理员账户', 'error')
        else:
            session.delete(user)
            session.commit()
            flash('用户已删除', 'success')
    session.close()
    return redirect(url_for('manage_users'))

def safe_float(val):
    try:
        if pd.isna(val):
            return 0.0
        val_str = str(val).strip()
        if not val_str or val_str == '/':
            return 0.0
        return float(val_str)
    except:
        return 0.0

@app.route('/cleared_hours')
def cleared_hours():
    sort_by = request.args.get('sort', 'reward_month')
    order = request.args.get('order', 'desc')
    search_query = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    session = Session()
    q = session.query(ClearedHoursRecord)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        q = q.join(Mechanic, ClearedHoursRecord.mechanic_id == Mechanic.id).filter(Mechanic.workshop_id == current_user.workshop_id)
    
    if search_query:
        q = q.filter((ClearedHoursRecord.name.contains(search_query)) | (ClearedHoursRecord.employee_id.contains(search_query)))
    
    if order == 'asc':
        q = q.order_by(asc(getattr(ClearedHoursRecord, sort_by)))
    else:
        q = q.order_by(desc(getattr(ClearedHoursRecord, sort_by)))
        
    # Pagination
    total_count = q.count()
    records = q.offset((page - 1) * per_page).limit(per_page).all()
    
    total_pages = (total_count + per_page - 1) // per_page
    start_index = (page - 1) * per_page + 1
    end_index = min(page * per_page, total_count)
    
    session.close()
    return render_template('cleared_hours.html', records=records,
                           sort_by=sort_by, order=order, search_query=search_query,
                           page=page, per_page=per_page, total_pages=total_pages,
                           total_count=total_count, start_index=start_index, end_index=end_index)

@app.route('/fetch_cleared_hours', methods=['POST'])
def fetch_cleared_hours():
    session = Session()
    try:
        # 1. Find mechanics with deduction >= 6 (value <= -6 if stored as negative, but current_cycle_deduction is Magnitude)
        # Note: In database.py: current_cycle_deduction = Column(Float, default=0.0)
        # Logic in upload(): mechanic.current_cycle_deduction += deduction_mag
        # So it stores Positive Magnitude.
        # User requirement: "current cycle deduction >= 6 (i.e. value <= -6)".
        # Since I store Magnitude, I should check >= 6.
        
        candidates_q = session.query(Mechanic).filter(Mechanic.current_cycle_deduction >= 6)
        if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
            candidates_q = candidates_q.filter(Mechanic.workshop_id == current_user.workshop_id)
        candidates = candidates_q.all()
        
        count_added = 0
        count_updated = 0
        
        for m in candidates:
            hist_count_before = len(m.rewards_history) if m.rewards_history else 0
            # Check if already in ClearedHoursRecord
            # How to identify? Employee ID.
            # But one person can be cleared multiple times?
            # Attachment 8 structure implies a summary per person ("Cumulative Clearing Count").
            # So we should find the existing record for this person and update it.
            
            record = session.query(ClearedHoursRecord).filter_by(employee_id=m.employee_id).first()
            
            # Determine "New Cleared Hours" value
            # User: "New Cleared Hours = Cumulative Hours in Thousand Hours Management"
            # That is m.total_hours.
            
            new_val = m.total_hours
            
            if record:
                # Update existing
                # Should we add to existing 'new' or overwrite?
                # If we "Fetch", we are syncing current state.
                # If we assume this is the *only* clearing event since 2026 start, overwrite is fine.
                # But if they were cleared in Jan, and now March.
                # Ideally we should know if this specific event was processed.
                # But without event log, we might just overwrite "New" with current total.
                # IF the user has been accumulating hours since last clear, `m.total_hours` IS the new amount to clear.
                # So Overwrite `cleared_hours_new` = `m.total_hours` seems correct for a "Sync" logic.
                # However, if they were cleared before, `m.total_hours` should have been reset?
                # But we agreed NOT to reset `m.total_hours` automatically in this step.
                # So `m.total_hours` keeps growing.
                # If Jan: 500h (Cleared). We Fetch. Record: New=500.
                # If Feb: +100h. Total=600h. We Fetch. Record: New=600.
                # This implies "Total Cleared Post-2026" is 600.
                # This seems consistent with "Cumulative".
                
                record.cleared_hours_new = new_val
                record.total_cleared_hours = record.cleared_hours_base + new_val
                
                # Update deduction details?
                # User says "Fetch... info... include in management".
                # Presumably we should update the "Current Cycle" info too.
                # Cycle Deduction
                record.cycle_deduction = m.current_cycle_deduction
                
                # Details
                issues = m.issues
                details_txt = "\n".join([i.problem for i in issues])
                record.deduction_details = details_txt
                
                # Reward Month/Cycle?
                # Maybe update to current?
                last_record = session.query(MonthlyRecord).order_by(MonthlyRecord.month.desc()).first()
                current_month = datetime.now().strftime('%Y.%m')
                end_month = last_record.month.replace('-', '.') if last_record else current_month
                cycle_str = f"2025.12.26-{end_month}.25"
                
                record.reward_month = current_month
                record.reward_cycle = cycle_str
                
                count_updated += 1
                
            else:
                # Create New
                # Need basic info (Name, Team, etc.)
                # Depot? mechanic.team might contain it or just use "Unknown"?
                # Attachment 8 has "Depot". Mechanic table doesn't have explicit Depot, just Team.
                # We can leave Depot blank or infer.
                
                last_record = session.query(MonthlyRecord).order_by(MonthlyRecord.month.desc()).first()
                current_month = datetime.now().strftime('%Y.%m')
                end_month = last_record.month.replace('-', '.') if last_record else current_month
                cycle_str = f"2025.12.26-{end_month}.25"
                
                issues = m.issues
                details_txt = "\n".join([i.problem for i in issues])
                
                record = ClearedHoursRecord(
                    mechanic_id=m.id,
                    depot="", # Placeholder
                    name=m.name,
                    employee_id=m.employee_id,
                    team=m.team,
                    activity_cumulative_hours=m.total_hours, # "Activity Cumulative" ~ Total Hours?
                    # Or is it History * 1000 + Total?
                    # In Mechanic list we display "Activity Cumulative" as History*1000 + Base.
                    # Let's use that logic if possible.
                    # (len(m.rewards_history) * 1000) + m.base_hours
                    # Actually, Attachment 8 has this column.
                    
                    reward_month=current_month,
                    reward_cycle=cycle_str,
                    cycle_deduction=m.current_cycle_deduction,
                    deduction_details=details_txt,
                    clearing_count=1, # First time in this list
                    
                    cleared_hours_base=0.0, # User Requirement: Base=0 for new
                    cleared_hours_new=new_val,
                    total_cleared_hours=new_val, # 0 + New
                    
                    remarks="自动获取"
                )
                
                # Calculate Activity Cumulative for new record
                hist_count = len(m.rewards_history)
                # If we assume `total_hours` is the current running count.
                # "Activity Cumulative" usually means Lifetime Total.
                # Let's use (History * 1000) + Total Hours.
                record.activity_cumulative_hours = (hist_count * 1000) + m.total_hours
                
                session.add(record)
                count_added += 1
        
        write_operation_log(
            "获取清零工时数据",
            json.dumps({"added": count_added, "updated": count_updated}, ensure_ascii=False),
            session=session
        )
        session.commit()
        flash(f'成功获取数据: 新增 {count_added} 条, 更新 {count_updated} 条', 'success')
        
    except Exception as e:
        session.rollback()
        flash(f'获取失败: {e}', 'error')
    finally:
        session.close()
        
    return redirect(url_for('cleared_hours'))

@app.route('/export/cleared_hours')
def export_cleared_hours():
    session = Session()
    records_q = session.query(ClearedHoursRecord)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        records_q = records_q.join(Mechanic, ClearedHoursRecord.mechanic_id == Mechanic.id).filter(Mechanic.workshop_id == current_user.workshop_id)
    records = records_q.all()
    
    template_path = os.path.join('excel_templates', '附件8 模板.xlsx')
    output_path = os.path.join('excel_templates', 'generated_cleared_hours.xlsx')
    shutil.copy(template_path, output_path)
    
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Template Check:
    # Row 1: Title?
    # Row 2: Header?
    # Data starts Row 3.
    # Let's assume header is row 2.
    row_idx = 3
    
    # Border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    i = 1
    for r in records:
        # Determine how many lines in deduction details
        details = r.deduction_details.split('\n') if r.deduction_details else []
        if not details:
            details = [""]
            
        # We need to write multiple rows if details > 1
        start_row = row_idx
        end_row = row_idx + len(details) - 1
        
        # Merge cells first (if needed)
        if len(details) > 1:
            cols_to_merge = [1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13]
            for col in cols_to_merge:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

        # Write data
        for idx, detail in enumerate(details):
            current_row = start_row + idx
            
            # Write Detail in Column J (10)
            c = ws.cell(row=current_row, column=10, value=detail)
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            
            # If it's the first row, write other columns
            if idx == 0:
                vals = [
                    (1, i),
                    (2, r.depot),
                    (3, r.name),
                    (4, r.employee_id),
                    (5, r.team),
                    (6, r.activity_cumulative_hours),
                    (7, r.reward_month),
                    (8, r.reward_cycle),
                    (9, r.cycle_deduction),
                    # 10 is detail
                    (11, r.clearing_count),
                    (12, r.total_cleared_hours), 
                    (13, r.remarks)
                ]
                for col, val in vals:
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                # Apply border to merged cells (OpenPyXL requires border on all cells in range or just top-left?)
                # For merged cells, border should be set on the top-left cell usually, 
                # but to be safe we can iterate.
                # Actually, if we merge first, setting value on top-left is enough.
                # But we need borders.
                for col in range(1, 14):
                    if col != 10:
                        c = ws.cell(row=current_row, column=col)
                        c.border = thin_border

        row_idx = end_row + 1
        i += 1
        
    wb.save(output_path)
    session.close()
    return send_file(output_path, as_attachment=True, download_name="工时清零汇总表.xlsx")

@app.route('/data_management')
@login_required
def data_management():
    # Overview of data types
    return render_template('data_management.html')

@app.route('/data/reset_annual', methods=['POST'])
@login_required
def reset_annual_flags():
    session = Session()
    try:
        # Update all existing issues: set include_in_annual = False
        # This marks them as "Past Year"
        # This is a bulk update
        session.query(Issue).update({Issue.include_in_annual: False}, synchronize_session=False)
        session.commit()
        
        # Log the operation
        write_operation_log(session, current_user.id if current_user.is_authenticated else 'System', 
                           '重置年度积分', '将所有现有问题标记为往年数据（不纳入年度积分统计）')
        
        flash('已成功重置年度积分统计（所有现有问题标记为往年数据）', 'success')
    except Exception as e:
        session.rollback()
        flash(f'重置失败: {e}', 'error')
    finally:
        session.close()
    return redirect(url_for('report_management'))

@app.route('/report_management')
@login_required
def report_management():
    reward_month = request.args.get('reward_month', datetime.now().strftime('%Y-%m'))
    return render_template('report_management.html', reward_month=reward_month)

@app.route('/data/monthly_hours')
@login_required
def data_monthly_hours():
    # List Monthly Records
    page = request.args.get('page', 1, type=int)
    search_month = request.args.get('month', '')
    search_name = request.args.get('name', '')
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in [20, 40, 60, 100]:
        per_page = 20
    session = Session()
    
    q = session.query(MonthlyRecord).join(Mechanic).options(joinedload(MonthlyRecord.mechanic))
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        q = q.filter(Mechanic.workshop_id == current_user.workshop_id)
    
    if search_month:
        q = q.filter(MonthlyRecord.month == search_month)
    if search_name:
        q = q.filter((Mechanic.name.contains(search_name)) | (Mechanic.employee_id.contains(search_name)))
        
    q = q.order_by(MonthlyRecord.month.desc())
    
    total = q.count()
    records = q.offset((page-1)*per_page).limit(per_page).all()
    
    # Get all distinct months for filter dropdown
    months_q = session.query(MonthlyRecord.month).join(Mechanic)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        months_q = months_q.filter(Mechanic.workshop_id == current_user.workshop_id)
    months = months_q.distinct().order_by(MonthlyRecord.month.desc()).all()
    months = [m[0] for m in months]
    
    session.close()
    
    total_pages = (total + per_page - 1) // per_page
    return render_template('data_list.html', title="月度工时数据", 
                           headers=['ID', '月份', '工号', '姓名', '工时'],
                           rows=[(r.id, r.month, r.mechanic.employee_id if r.mechanic else '', r.mechanic.name if r.mechanic else '', f"{r.hours:.2f}") for r in records],
                           page=page, total_pages=total_pages, endpoint='data_monthly_hours',
                           delete_endpoint='delete_monthly_hours',
                           edit_endpoint='edit_monthly_hours',
                           per_page=per_page,
                           filters={'month': months, 'current_month': search_month, 'name': search_name})

@app.route('/data/edit/monthly_hours/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_monthly_hours(id):
    session = Session()
    rec = session.query(MonthlyRecord).get(id)
    if not rec:
        session.close()
        flash('记录不存在', 'error')
        return redirect(url_for('data_monthly_hours'))
        
    if request.method == 'POST':
        try:
            old_hours = rec.hours
            new_hours = safe_float(request.form.get('hours'))
            
            # Update Mechanic total hours
            if rec.mechanic:
                rec.mechanic.total_hours = rec.mechanic.total_hours - old_hours + new_hours
                
            rec.hours = new_hours
            write_operation_log(
                "修改月度工时",
                json.dumps({
                    "record_id": rec.id,
                    "mechanic_id": rec.mechanic_id,
                    "old_hours": old_hours,
                    "new_hours": new_hours
                }, ensure_ascii=False),
                session=session
            )
            session.commit()
            flash('修改成功', 'success')
            return redirect(url_for('data_monthly_hours'))
        except Exception as e:
            session.rollback()
            flash(f'修改失败: {e}', 'error')
            
    return render_template('data_edit.html', title="修改月度工时", 
                           fields=[
                               {'name': 'hours', 'label': '工时', 'value': rec.hours, 'type': 'number', 'step': '0.01'}
                           ],
                           action=url_for('edit_monthly_hours', id=id))

@app.route('/data/monthly_issues')
@login_required
def data_monthly_issues():
    # List Issues
    page = request.args.get('page', 1, type=int)
    search_month = request.args.get('month', '')
    search_name = request.args.get('name', '')
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in [20, 40, 60, 100]:
        per_page = 20
    session = Session()
    q = session.query(Issue).join(Mechanic).options(joinedload(Issue.mechanic)).order_by(Issue.date.desc())
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        q = q.filter(Mechanic.workshop_id == current_user.workshop_id)
    if search_name:
        q = q.filter((Mechanic.name.contains(search_name)) | (Mechanic.employee_id.contains(search_name)))
    
    records_all = q.all()
    if search_month:
        records_all = [r for r in records_all if is_in_month_common(r.date, search_month)]
    
    total = len(records_all)
    start = (page - 1) * per_page
    end = start + per_page
    records = records_all[start:end]
    
    months_q = session.query(MonthlyRecord.month).join(Mechanic)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        months_q = months_q.filter(Mechanic.workshop_id == current_user.workshop_id)
    months = months_q.distinct().order_by(MonthlyRecord.month.desc()).all()
    months = [m[0] for m in months]
    
    session.close()
    
    # Format Date Helper
    def fmt_date(d):
        try:
            # Check if Excel Serial Date (e.g. 45000)
            if d and str(d).replace('.','').isdigit() and float(d) > 30000:
                dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(float(d)) - 2)
                return dt.strftime('%Y-%m-%d')
            return d
        except:
            return d

    total_pages = (total + per_page - 1) // per_page
    return render_template('data_list.html', title="月度问题数据",
                           headers=['ID', '日期', '工号', '姓名', '问题', '来源', '条款', '明细', '总计', '状态'],
                           rows=[(r.id, fmt_date(r.date), r.mechanic.employee_id if r.mechanic else '', r.mechanic.name if r.mechanic else '', r.problem, r.source, r.clause, r.detail, r.total_deduction, getattr(r, 'status', '未结算')) for r in records],
                           page=page, total_pages=total_pages, endpoint='data_monthly_issues',
                           delete_endpoint='delete_issue',
                           edit_endpoint='edit_issue',
                           per_page=per_page,
                           filters={'month': months, 'current_month': search_month, 'name': search_name})

@app.route('/data/edit/issue/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_issue(id):
    session = Session()
    rec = session.query(Issue).get(id)
    if not rec:
        session.close()
        flash('记录不存在', 'error')
        return redirect(url_for('data_monthly_issues'))
        
    if request.method == 'POST':
        try:
            old_detail = rec.detail or 0
            
            rec.date = request.form.get('date')
            rec.problem = request.form.get('problem')
            rec.source = request.form.get('source')
            rec.clause = request.form.get('clause')
            rec.detail = safe_float(request.form.get('detail'))
            rec.total_deduction = safe_float(request.form.get('total_deduction'))
            
            # Update Mechanic Deduction Logic?
            # User wants to manage data. If we change deduction amount, we should update mechanic's current cycle deduction.
            if rec.mechanic:
                diff = abs(rec.detail) - abs(old_detail)
                rec.mechanic.current_cycle_deduction += diff
                if rec.mechanic.current_cycle_deduction < 0: rec.mechanic.current_cycle_deduction = 0
            
            session.commit()
            flash('修改成功', 'success')
            return redirect(url_for('data_monthly_issues'))
        except Exception as e:
            session.rollback()
            flash(f'修改失败: {e}', 'error')
            
    return render_template('data_edit.html', title="修改问题记录", 
                           fields=[
                               {'name': 'date', 'label': '日期', 'value': rec.date, 'type': 'text'},
                               {'name': 'problem', 'label': '问题', 'value': rec.problem, 'type': 'text'},
                               {'name': 'source', 'label': '来源', 'value': rec.source, 'type': 'text'},
                               {'name': 'clause', 'label': '条款', 'value': rec.clause, 'type': 'text'},
                               {'name': 'detail', 'label': '明细(扣分)', 'value': rec.detail, 'type': 'number', 'step': '0.01'},
                               {'name': 'total_deduction', 'label': '总计', 'value': rec.total_deduction, 'type': 'number', 'step': '0.01'}
                           ],
                           action=url_for('edit_issue', id=id))

@app.route('/data/att6')
@login_required
def data_att6():
    # List Attachment 6 Data
    page = request.args.get('page', 1, type=int)
    search_month = request.args.get('month', '')
    search_name = request.args.get('name', '')
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in [20, 40, 60, 100]:
        per_page = 20
    session = Session()
    q = session.query(Attachment6Data).join(Mechanic, Attachment6Data.employee_id == Mechanic.employee_id)
    
    # Workshop Filter
    if current_user.role == ROLE_WORKSHOP:
        if current_user.workshop_id:
            q = q.filter(Mechanic.workshop_id == current_user.workshop_id)
    
    if search_month:
        # Filter by Reward Date (YYYY-MM or YYYY.MM)
        # Assuming format in DB is varied
        prefix_dash = search_month + '%'
        prefix_dot = search_month.replace('-', '.') + '%'
        q = q.filter(or_(Attachment6Data.reward_date.like(prefix_dash), Attachment6Data.reward_date.like(prefix_dot)))
        
    if search_name:
        q = q.filter(or_(Attachment6Data.name.contains(search_name), Attachment6Data.employee_id.contains(search_name)))
        
    q = q.order_by(Attachment6Data.name.asc(), Attachment6Data.reward_date.desc())
        
    total = q.count()
    records = q.offset((page-1)*per_page).limit(per_page).all()
    
    # Distinct months
    # Fetch distinct reward_date prefixes?
    # Or just use MonthlyRecord months as proxy?
    # Or fetch all reward_dates and format them.
    # Att 6 usually has cleaner dates?
    dates_q = session.query(Attachment6Data.reward_date).join(Mechanic, Attachment6Data.employee_id == Mechanic.employee_id)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        dates_q = dates_q.filter(Mechanic.workshop_id == current_user.workshop_id)
    dates = dates_q.distinct().all()
    # Parse and unique YYYY-MM
    months = set()
    for d in dates:
        if d[0]:
            try:
                # Format: YYYY.MM or YYYY-MM
                s = str(d[0]).replace('.', '-')
                if len(s) >= 7:
                    months.add(s[:7])
            except: pass
    months = sorted(list(months), reverse=True)
    
    # Helper to format data
    def fmt_val(v):
        try:
            return f"{float(v):.2f}"
        except:
            return v
            
    def fmt_date(d):
        try:
            # Try to parse if it contains time
            if ' ' in str(d):
                return str(d).split(' ')[0]
            # If it's Excel numeric
            if str(d).replace('.','').isdigit() and float(d) > 30000:
                dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(float(d)) - 2)
                return dt.strftime('%Y-%m-%d')
            return d
        except:
            return d

    def fmt_cycle(c):
        if not c: return c
        try:
            # Expected format: YYYY.MM.DD-YYYY.MM.DD or YYYY.MM-YYYY.MM
            # Target: YYYY年MM月-YYYY年MM月
            # Split by '-'
            if '-' in str(c):
                parts = str(c).split('-')
                new_parts = []
                for p in parts:
                    p = p.strip()
                    # Try parsing date
                    # If it has dots
                    if '.' in p:
                        subparts = p.split('.')
                        if len(subparts) >= 2:
                            # Year and Month
                            year = subparts[0]
                            month = int(subparts[1])
                            new_parts.append(f"{year}年{month}月")
                        else:
                            new_parts.append(p)
                    # If already has Chinese
                    elif '年' in p and '月' in p:
                        new_parts.append(p)
                    else:
                        new_parts.append(p)
                return "-".join(new_parts)
            return c
        except:
            return c

    def normalize_month(val):
        if not val:
            return None
        s = str(val).strip()
        if not s:
            return None
        s = s.replace('年', '-').replace('月', '').replace('.', '-').replace('/', '-')
        parts = s.split('-')
        if len(parts) >= 2:
            y = parts[0]
            m = parts[1]
            if len(m) == 1:
                m = f"0{m}"
            return f"{y}-{m}"
        if len(s) == 6 and s.isdigit():
            return f"{s[:4]}-{s[4:]}"
        return None

    records_employee_ids = [r.employee_id for r in records if r.employee_id]
    employee_to_mid = {}
    if records_employee_ids:
        mech_rows = session.query(Mechanic.id, Mechanic.employee_id).filter(Mechanic.employee_id.in_(records_employee_ids)).all()
        employee_to_mid = {r[1]: r[0] for r in mech_rows}

    mids = list({mid for mid in employee_to_mid.values() if mid})
    prev_extra_map = {}
    if mids:
        hist_rows = session.query(RewardHistory.mechanic_id, RewardHistory.reward_date, RewardHistory.extra_reward).filter(
            RewardHistory.mechanic_id.in_(mids)
        ).order_by(RewardHistory.mechanic_id.asc(), RewardHistory.reward_date.asc()).all()
        per_mid = {}
        for mid, rdate, extra in hist_rows:
            mkey = normalize_month(rdate) or str(rdate or '').strip()
            if not mkey:
                continue
            per_mid.setdefault(mid, []).append((mkey, float(extra or 0.0)))
        for mid, arr in per_mid.items():
            seen = {}
            for mkey, extra in arr:
                seen[mkey] = extra
            keys_sorted = sorted(seen.keys())
            prev = 0.0
            for k in keys_sorted:
                prev_extra_map[(mid, k)] = prev
                prev = float(seen[k] or 0.0)

    def parse_number(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        for p in ['上期：', '上期:', '上期']:
            if s.startswith(p):
                s = s[len(p):].strip()
        try:
            return float(s)
        except:
            return None

    total_pages = (total + per_page - 1) // per_page
    session.close()
    return render_template('data_list.html', title="附件6原始数据",
                           headers=['ID', '姓名', '工号', '乘务组别', '活动起累计工时', '本次满足奖励时间', '本次奖励周期', '周期累积扣分', '本次奖励周期奖励金额', '累计清零工时', '当前前结余工时', '过去已奖励次数', '是否触发连续奖励', '连续奖励情况', '本次连续额外奖励', '本月奖励金额'],
                           rows=[(r.id, r.name, r.employee_id, r.team, fmt_val(r.activity_cumulative_hours), fmt_date(r.reward_date), fmt_cycle(r.reward_cycle), r.cycle_deduction, r.reward_amount, r.cleared_hours, fmt_val(r.balance_hours), r.past_reward_count, r.is_consecutive, fmt_val(prev_extra_map.get((employee_to_mid.get(r.employee_id), normalize_month(r.reward_date) or str(r.reward_date or '').strip()), (parse_number(r.consecutive_info) or 0.0))), r.extra_reward, r.total_amount) for r in records],
                           page=page, total_pages=total_pages, endpoint='data_att6',
                           delete_endpoint='delete_att6',
                           edit_endpoint='edit_att6',
                           per_page=per_page,
                           filters={'month': months, 'current_month': search_month, 'name': search_name})

@app.route('/data/route_hours')
@login_required
def data_route_hours():
    page = request.args.get('page', 1, type=int)
    search_train = request.args.get('name', '')
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in [20, 40, 60, 100]:
        per_page = 20

    session = Session()
    try:
        if session.query(RouteHoursData).count() == 0:
            path = os.path.join('使用文件', '交路工时信息.xlsx')
            if os.path.exists(path):
                import openpyxl
                wb = openpyxl.load_workbook(path)
                ws = wb.active

                def norm(s):
                    if s is None:
                        return ''
                    return re.sub(r'\s+', '', str(s)).strip()

                header_row = None
                header_map = {}
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    vals = [norm(v) for v in row]
                    if any('车次' in v for v in vals):
                        header_row = i
                        for idx, v in enumerate(vals):
                            if v:
                                header_map[v] = idx
                        break

                def find_idx(keys):
                    for k in header_map.keys():
                        if any(x in k for x in keys):
                            return header_map[k]
                    return None

                idx_train = find_idx(['车次', '车次号', '车次/'])
                idx_pre = find_idx(['出乘前作业时间'])
                idx_en = find_idx(['途中运行时间'])
                idx_post = find_idx(['入库后作业时间'])
                idx_remark = find_idx(['备注'])

                if header_row and idx_train is not None:
                    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                        train_no = str(row[idx_train]).strip() if idx_train < len(row) and row[idx_train] is not None else ''
                        if not train_no:
                            continue
                        rec = RouteHoursData(
                            train_no=train_no,
                            pre_work_hours=safe_float(row[idx_pre]) if idx_pre is not None and idx_pre < len(row) else 0.0,
                            enroute_hours=safe_float(row[idx_en]) if idx_en is not None and idx_en < len(row) else 0.0,
                            post_work_hours=safe_float(row[idx_post]) if idx_post is not None and idx_post < len(row) else 0.0,
                            remarks=str(row[idx_remark]).strip() if idx_remark is not None and idx_remark < len(row) and row[idx_remark] is not None else None
                        )
                        session.add(rec)
                    session.commit()

        q = session.query(RouteHoursData)
        if search_train:
            q = q.filter(RouteHoursData.train_no.contains(search_train))
        q = q.order_by(RouteHoursData.train_no.asc())

        total = q.count()
        records = q.offset((page - 1) * per_page).limit(per_page).all()
        total_pages = (total + per_page - 1) // per_page
        return render_template(
            'data_list.html',
            title='交路工时数据',
            headers=['ID', '车次', '出乘前作业时间', '途中运行时间', '入库后作业时间', '备注'],
            rows=[(r.id, r.train_no, f"{float(r.pre_work_hours or 0.0):.2f}", f"{float(r.enroute_hours or 0.0):.2f}", f"{float(r.post_work_hours or 0.0):.2f}", r.remarks or '') for r in records],
            page=page,
            total_pages=total_pages,
            endpoint='data_route_hours',
            delete_endpoint='delete_route_hours',
            edit_endpoint='edit_route_hours',
            create_endpoint='create_route_hours',
            per_page=per_page,
            filters={'name': search_train, 'name_label': '车次', 'name_placeholder': '输入车次'}
        )
    finally:
        session.close()

@app.route('/data/route_hours/new', methods=['GET', 'POST'])
@login_required
def create_route_hours():
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('data_route_hours'))

    if request.method == 'POST':
        train_no = (request.form.get('train_no') or '').strip()
        if not train_no:
            flash('车次不能为空', 'error')
            return redirect(url_for('create_route_hours'))
        session = Session()
        try:
            rec = RouteHoursData(
                train_no=train_no,
                pre_work_hours=safe_float(request.form.get('pre_work_hours')),
                enroute_hours=safe_float(request.form.get('enroute_hours')),
                post_work_hours=safe_float(request.form.get('post_work_hours')),
                remarks=request.form.get('remarks')
            )
            session.add(rec)
            session.commit()
            flash('新增成功', 'success')
            return redirect(url_for('data_route_hours'))
        except Exception as e:
            session.rollback()
            flash(f'新增失败: {e}', 'error')
            return redirect(url_for('create_route_hours'))
        finally:
            session.close()

    fields = [
        {'name': 'train_no', 'label': '车次', 'value': '', 'type': 'text'},
        {'name': 'pre_work_hours', 'label': '出乘前作业时间', 'value': 0, 'type': 'number', 'step': '0.01'},
        {'name': 'enroute_hours', 'label': '途中运行时间', 'value': 0, 'type': 'number', 'step': '0.01'},
        {'name': 'post_work_hours', 'label': '入库后作业时间', 'value': 0, 'type': 'number', 'step': '0.01'},
        {'name': 'remarks', 'label': '备注', 'value': '', 'type': 'text'}
    ]
    return render_template('data_edit.html', title='新增交路工时', fields=fields, action=url_for('create_route_hours'))

@app.route('/data/edit/route_hours/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_route_hours(id):
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('data_route_hours'))

    session = Session()
    rec = session.query(RouteHoursData).get(id)
    if not rec:
        session.close()
        flash('记录不存在', 'error')
        return redirect(url_for('data_route_hours'))

    if request.method == 'POST':
        try:
            train_no = (request.form.get('train_no') or '').strip()
            if not train_no:
                flash('车次不能为空', 'error')
                return redirect(url_for('edit_route_hours', id=id))
            rec.train_no = train_no
            rec.pre_work_hours = safe_float(request.form.get('pre_work_hours'))
            rec.enroute_hours = safe_float(request.form.get('enroute_hours'))
            rec.post_work_hours = safe_float(request.form.get('post_work_hours'))
            rec.remarks = request.form.get('remarks')
            session.commit()
            flash('修改成功', 'success')
            return redirect(url_for('data_route_hours'))
        except Exception as e:
            session.rollback()
            flash(f'修改失败: {e}', 'error')
            return redirect(url_for('edit_route_hours', id=id))
        finally:
            session.close()

    fields = [
        {'name': 'train_no', 'label': '车次', 'value': rec.train_no or '', 'type': 'text'},
        {'name': 'pre_work_hours', 'label': '出乘前作业时间', 'value': rec.pre_work_hours or 0, 'type': 'number', 'step': '0.01'},
        {'name': 'enroute_hours', 'label': '途中运行时间', 'value': rec.enroute_hours or 0, 'type': 'number', 'step': '0.01'},
        {'name': 'post_work_hours', 'label': '入库后作业时间', 'value': rec.post_work_hours or 0, 'type': 'number', 'step': '0.01'},
        {'name': 'remarks', 'label': '备注', 'value': rec.remarks or '', 'type': 'text'}
    ]
    session.close()
    return render_template('data_edit.html', title='修改交路工时', fields=fields, action=url_for('edit_route_hours', id=id))

@app.route('/data/delete/route_hours/<int:id>')
@login_required
def delete_route_hours(id):
    if current_user.role != ROLE_SECTION:
        flash('权限不足', 'error')
        return redirect(url_for('data_route_hours'))

    session = Session()
    rec = session.query(RouteHoursData).get(id)
    if rec:
        session.delete(rec)
        session.commit()
        flash('删除成功', 'success')
    session.close()
    return redirect(url_for('data_route_hours'))

@app.route('/data/delete_batch/monthly_hours', methods=['POST'])
@login_required
def delete_batch_monthly_hours():
    month = request.form.get('month')
    if not month:
        flash('请选择要删除的月份', 'error')
        return redirect(url_for('data_monthly_hours'))
        
    session = Session()
    try:
        # Find records
        records = session.query(MonthlyRecord).filter_by(month=month).all()
        count = 0
        for rec in records:
            if rec.mechanic:
                rec.mechanic.total_hours -= rec.hours
            session.delete(rec)
            count += 1
        
        write_operation_log(
            "批量删除月度工时",
            json.dumps({"month": month, "count": count}, ensure_ascii=False),
            session=session
        )
        session.commit()
        flash(f'成功删除 {month} 月份的 {count} 条工时记录，累计工时已更新', 'success')
    except Exception as e:
        session.rollback()
        flash(f'删除失败: {e}', 'error')
    finally:
        session.close()
        
    return redirect(url_for('data_monthly_hours'))

@app.route('/data/delete_batch/monthly_issues', methods=['POST'])
@login_required
def delete_batch_monthly_issues():
    # User might select a "Month" (YYYY-MM), but issues have YYYY-MM-DD or numeric date.
    # We rely on string matching or parsed date logic.
    month = request.form.get('month') # YYYY-MM
    if not month:
        flash('请选择要删除的月份', 'error')
        return redirect(url_for('data_monthly_issues'))
        
    session = Session()
    try:
        # Strategy: Iterate all issues, parse date, check if matches month
        # This is slow but safer given variable date formats.
        # Or try SQL LIKE if format is consistent.
        # Most imported dates are YYYY-MM-DD.
        # Let's try matching string start first, then fallback to parsing.
        
        # We need to find issues where date falls in month.
        # Helper to check
        def is_in_month(date_str, target_month):
            if not date_str: return False
            try:
                # Format: YYYY-MM
                if str(date_str).startswith(target_month):
                    return True
                # Format: YYYY/MM
                target_slash = target_month.replace('-', '/')
                if str(date_str).startswith(target_slash):
                    return True
                # Format: Excel serial
                if str(date_str).replace('.','').isdigit() and float(date_str) > 30000:
                    dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(float(date_str)) - 2)
                    return dt.strftime('%Y-%m') == target_month
                
                # Format: YYYY.MM
                target_dot = target_month.replace('-', '.')
                if str(date_str).startswith(target_dot):
                    return True
                    
                return False
            except:
                return False

        all_issues = session.query(Issue).all()
        to_delete = []
        for issue in all_issues:
            if is_in_month(issue.date, month):
                to_delete.append(issue)
        
        count = 0
        for issue in to_delete:
            if issue.mechanic:
                issue.mechanic.current_cycle_deduction -= abs(issue.detail or 0)
                if issue.mechanic.current_cycle_deduction < 0: issue.mechanic.current_cycle_deduction = 0
            session.delete(issue)
            count += 1
        
        write_operation_log(
            "批量删除月度问题",
            json.dumps({"month": month, "count": count}, ensure_ascii=False),
            session=session
        )
        session.commit()
        flash(f'成功删除 {month} 月份的 {count} 条问题记录，周期扣分已更新', 'success')
        
    except Exception as e:
        session.rollback()
        flash(f'删除失败: {e}', 'error')
    finally:
        session.close()
        
    return redirect(url_for('data_monthly_issues'))

@app.route('/data/edit/att6/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_att6(id):
    session = Session()
    rec = session.query(Attachment6Data).get(id)
    if not rec:
        session.close()
        flash('记录不存在', 'error')
        return redirect(url_for('data_att6'))
        
    if request.method == 'POST':
        try:
            # User said: "Data here does not link with other data for now". So just update record.
            rec.name = request.form.get('name')
            rec.employee_id = request.form.get('employee_id')
            rec.team = request.form.get('team')
            rec.activity_cumulative_hours = safe_float(request.form.get('activity_cumulative_hours'))
            rec.reward_date = request.form.get('reward_date')
            rec.reward_cycle = request.form.get('reward_cycle')
            rec.cycle_deduction = safe_float(request.form.get('cycle_deduction'))
            rec.reward_amount = safe_float(request.form.get('reward_amount'))
            rec.cleared_hours = safe_float(request.form.get('cleared_hours'))
            rec.balance_hours = safe_float(request.form.get('balance_hours'))
            rec.past_reward_count = int(safe_float(request.form.get('past_reward_count')))
            rec.is_consecutive = request.form.get('is_consecutive')
            rec.consecutive_info = request.form.get('consecutive_info')
            rec.extra_reward = safe_float(request.form.get('extra_reward'))
            rec.total_amount = safe_float(request.form.get('total_amount'))
            
            session.commit()
            flash('修改成功', 'success')
            return redirect(url_for('data_att6'))
        except Exception as e:
            session.rollback()
            flash(f'修改失败: {e}', 'error')
            
    return render_template('data_edit.html', title="修改附件6数据", 
                           fields=[
                               {'name': 'name', 'label': '姓名', 'value': rec.name, 'type': 'text'},
                               {'name': 'employee_id', 'label': '工号', 'value': rec.employee_id, 'type': 'text'},
                               {'name': 'team', 'label': '乘务组别', 'value': rec.team, 'type': 'text'},
                               {'name': 'activity_cumulative_hours', 'label': '活动起累计工时', 'value': rec.activity_cumulative_hours, 'type': 'number', 'step': '0.01'},
                               {'name': 'reward_date', 'label': '本次满足奖励时间', 'value': rec.reward_date, 'type': 'text'},
                               {'name': 'reward_cycle', 'label': '本次奖励周期', 'value': rec.reward_cycle, 'type': 'text'},
                               {'name': 'cycle_deduction', 'label': '周期扣分', 'value': rec.cycle_deduction, 'type': 'number', 'step': '0.01'},
                               {'name': 'reward_amount', 'label': '奖励金额', 'value': rec.reward_amount, 'type': 'number', 'step': '0.01'},
                               {'name': 'cleared_hours', 'label': '累计清零工时', 'value': rec.cleared_hours, 'type': 'number', 'step': '0.01'},
                               {'name': 'balance_hours', 'label': '当前前结余工时', 'value': rec.balance_hours, 'type': 'number', 'step': '0.01'},
                               {'name': 'past_reward_count', 'label': '过去已奖励次数', 'value': rec.past_reward_count, 'type': 'number'},
                               {'name': 'is_consecutive', 'label': '触发连续奖励', 'value': rec.is_consecutive, 'type': 'text'},
                               {'name': 'consecutive_info', 'label': '连续奖励情况', 'value': rec.consecutive_info, 'type': 'text'},
                               {'name': 'extra_reward', 'label': '本次连续额外奖励', 'value': rec.extra_reward, 'type': 'number', 'step': '0.01'},
                               {'name': 'total_amount', 'label': '本月奖励金额', 'value': rec.total_amount, 'type': 'number', 'step': '0.01'}
                           ],
                           action=url_for('edit_att6', id=id))

@app.route('/data/att1')
@login_required
def data_att1():
    # List Attachment 1 Data (Deduction Rules)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in [20, 40, 60, 100]:
        per_page = 20
    session = Session()
    q = session.query(Attachment1Data)
    total = q.count()
    records = q.offset((page-1)*per_page).limit(per_page).all()
    session.close()
    
    total_pages = (total + per_page - 1) // per_page
    return render_template('data_list.html', title="附件1扣分清单",
                           headers=['ID', '扣分大类', '具体项目', '详细描述', '分值'],
                           rows=[(r.id, r.source, r.clause, r.detail, r.score) for r in records],
                           page=page, total_pages=total_pages, endpoint='data_att1',
                           delete_endpoint='delete_att1',
                           per_page=per_page)

@app.route('/data/delete/monthly_hours/<int:id>')
@login_required
def delete_monthly_hours(id):
    session = Session()
    rec = session.query(MonthlyRecord).get(id)
    if rec:
        # Update mechanic total hours
        if rec.mechanic:
            rec.mechanic.total_hours -= rec.hours
        write_operation_log(
            "删除月度工时",
            json.dumps({"record_id": rec.id, "mechanic_id": rec.mechanic_id, "hours": rec.hours}, ensure_ascii=False),
            session=session
        )
        session.delete(rec)
        session.commit()
        flash('记录已删除，累计工时已更新', 'success')
    session.close()
    return redirect(url_for('data_monthly_hours'))

@app.route('/data/delete/issue/<int:id>')
@login_required
def delete_issue(id):
    session = Session()
    rec = session.query(Issue).get(id)
    if rec:
        # Update mechanic deduction?
        # If we delete an issue, we should reduce the current_cycle_deduction.
        if rec.mechanic:
            rec.mechanic.current_cycle_deduction -= abs(rec.detail or 0)
            if rec.mechanic.current_cycle_deduction < 0: rec.mechanic.current_cycle_deduction = 0
        write_operation_log(
            "删除问题记录",
            json.dumps({"issue_id": rec.id, "mechanic_id": rec.mechanic_id, "deduction": rec.detail, "date": rec.date}, ensure_ascii=False),
            session=session
        )
        session.delete(rec)
        session.commit()
        flash('问题记录已删除，当前周期扣分已更新', 'success')
    session.close()
    return redirect(url_for('data_monthly_issues'))

@app.route('/data/delete/att6/<int:id>')
@login_required
def delete_att6(id):
    session = Session()
    rec = session.query(Attachment6Data).get(id)
    if rec:
        # Also try to delete from RewardHistory to maintain consistency
        # Match by Employee ID and Cycle (and maybe Date)
        mechanic = session.query(Mechanic).filter_by(employee_id=rec.employee_id).first()
        if mechanic:
            # Find matching history
            # Note: dates might be slightly different strings if parsed differently, but let's try exact match
            history = session.query(RewardHistory).filter_by(
                mechanic_id=mechanic.id,
                reward_cycle=rec.reward_cycle
            ).first()
            if history:
                session.delete(history)
        
        session.delete(rec)
        session.commit()
        flash('附件6记录及对应奖励历史已删除', 'success')
    session.close()
    return redirect(url_for('data_att6'))

@app.route('/data/delete/att1/<int:id>')
@login_required
def delete_att1(id):
    session = Session()
    rec = session.query(Attachment1Data).get(id)
    if rec:
        session.delete(rec)
        session.commit()
        flash('扣分规则已删除', 'success')
    session.close()
    return redirect(url_for('data_att1'))

@app.route('/download/template/<type>')
def download_template(type):
    filename = ""
    if type == 'hours':
        filename = '月度工时模板.xlsx'
    elif type == 'issues':
        filename = '月度问题模板.xlsx'
    
    if filename:
        path = os.path.join('excel_templates', filename)
        if os.path.exists(path):
            return send_file(path, as_attachment=True)
    
    flash('模板文件不存在', 'error')
    return redirect(url_for('index'))

@app.route('/')
@login_required
def index():
    session = Session()
    mechanic_query = session.query(Mechanic).filter(Mechanic.status == '在岗')
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        mechanic_query = mechanic_query.filter(Mechanic.workshop_id == current_user.workshop_id)
    total = mechanic_query.count()
    over_1000 = mechanic_query.filter(Mechanic.total_hours >= 1000).count()
    backup_count = mechanic_query.filter(Mechanic.team == '后备').count()
    regular_count = total - backup_count
    
    last_record = session.query(MonthlyRecord).order_by(MonthlyRecord.month.desc()).first()
    if last_record:
        count = session.query(MonthlyRecord).filter_by(month=last_record.month).count()
        if count > 0:
            last_month_imported = last_record.month
        else:
            last_month_imported = "无数据"
    else:
        last_month_imported = "无数据"

    # Filter Params
    now = datetime.now()
    filter_year = request.args.get('year', now.year, type=int)
    filter_start_month = request.args.get('start_month', f"{now.year}-01")
    filter_end_month = request.args.get('end_month', f"{now.year}-12")
    trend_month = request.args.get('trend_month', last_month_imported if last_month_imported != "无数据" else now.strftime('%Y-%m'))
    trend_name = request.args.get('trend_name', '')
    trend_team = request.args.get('trend_team', '')

    hours_month = request.args.get('hours_month', trend_month)
    hours_mechanic = request.args.get('hours_mechanic', trend_name)
    hours_teams = [t for t in request.args.getlist('hours_team') if t]

    mechanics = mechanic_query.options(joinedload(Mechanic.issues), joinedload(Mechanic.rewards_history)).all()
    mechanic_names = sorted({m.name for m in mechanics if m.name})
    mechanic_teams = sorted({m.team for m in mechanics if m.team})
    
    score_labels = ["12", "11", "9-10", "7-8", "0-6"]
    score_counts = [0, 0, 0, 0, 0]
    deduction_labels = ["0", "0-1", "1-3", "3-6", ">=6"]
    deduction_counts = [0, 0, 0, 0, 0]
    reward_labels = ["1200", "1000", "600", "200", "0"]
    reward_counts = [0, 0, 0, 0, 0]
    hours_labels = ["<500", "500-800", "800-1000", "1000-1200", ">=1200"]
    hours_counts = [0, 0, 0, 0, 0]

    # 1. Annual Score (Filter by Year)
    # 2. Deduction Distribution (Filter by Month Range)
    # 3. Hours Distribution (Current Total)
    
    for m in mechanics:
        # --- Annual Score ---
        # Filter issues by Year
        annual_issues = []
        for i in m.issues:
            if getattr(i, 'include_in_annual', True) and i.date and str(i.date).startswith(str(filter_year)):
                annual_issues.append(i)
        
        deduction_annual = abs(sum(i.detail or 0 for i in annual_issues))
        current_score = 12 - deduction_annual

        if current_score >= 12: score_counts[0] += 1
        elif current_score >= 11: score_counts[1] += 1
        elif current_score >= 9: score_counts[2] += 1
        elif current_score >= 7: score_counts[3] += 1
        else: score_counts[4] += 1

        # --- Deduction Distribution (Range) ---
        range_issues = []
        for i in m.issues:
            if not i.date: continue
            # Date format assumed YYYY-MM-DD
            d_str = str(i.date)[:7] # YYYY-MM
            if filter_start_month <= d_str <= filter_end_month:
                range_issues.append(i)
        
        deduction_range = abs(sum(i.detail or 0 for i in range_issues))
        
        if deduction_range == 0: deduction_counts[0] += 1
        elif deduction_range <= 1: deduction_counts[1] += 1
        elif deduction_range < 3: deduction_counts[2] += 1
        elif deduction_range < 6: deduction_counts[3] += 1
        else: deduction_counts[4] += 1

        # --- Hours Distribution (Current) ---
        hours_val = m.total_hours or 0.0
        if hours_val < 500: hours_counts[0] += 1
        elif hours_val < 800: hours_counts[1] += 1
        elif hours_val < 1000: hours_counts[2] += 1
        elif hours_val < 1200: hours_counts[3] += 1
        else: hours_counts[4] += 1

    # --- Reward Distribution (Range) ---
    # Based on Attachment6Data (Actual Rewards)
    # Normalize dates for query
    s_dot = filter_start_month.replace('-', '.')
    e_dot = filter_end_month.replace('-', '.')
    
    # We fetch all and filter in python to handle mixed formats safely
    att6_query = session.query(Attachment6Data)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        att6_query = att6_query.join(Mechanic, Attachment6Data.employee_id == Mechanic.employee_id).filter(Mechanic.workshop_id == current_user.workshop_id)
    att6_all = att6_query.all()
    for r in att6_all:
        if not r.reward_date: continue
        d = str(r.reward_date).strip()
        # Normalize to YYYY-MM
        d_norm = d.replace('.', '-').replace('年', '-').replace('月', '')
        if len(d_norm) == 6: # YYYY-M -> YYYY-0M
            parts = d_norm.split('-')
            if len(parts) == 2 and len(parts[1]) == 1:
                d_norm = f"{parts[0]}-0{parts[1]}"
        
        # Check range
        # Only check YYYY-MM
        d_chk = d_norm[:7]
        if filter_start_month <= d_chk <= filter_end_month:
            amt = r.reward_amount or 0
            if amt >= 1200: reward_counts[0] += 1
            elif amt >= 1000: reward_counts[1] += 1
            elif amt >= 600: reward_counts[2] += 1
            elif amt >= 200: reward_counts[3] += 1
            else: reward_counts[4] += 1
    
    # Historical Trend (Range Filter)
    att6_records = att6_query.with_entities(Attachment6Data.reward_date, Attachment6Data.total_amount).all()
    history_map = {}
    
    for r in att6_records:
        r_date = r.reward_date
        if not r_date: continue
        try:
            norm_date = r_date.strip().replace('.', '-').replace('年', '-').replace('月', '')
            parts = norm_date.split('-')
            if len(parts) >= 2:
                y = parts[0]
                m = parts[1]
                if len(m) == 1: m = '0' + m
                norm_date = f"{y}-{m}"
            
            # Apply Range Filter to Trend
            if filter_start_month <= norm_date <= filter_end_month:
                if norm_date not in history_map:
                    history_map[norm_date] = {'amount': 0.0, 'count': 0}
                history_map[norm_date]['amount'] += (r.total_amount or 0)
                history_map[norm_date]['count'] += 1
        except: pass
        
    sorted_months = sorted(history_map.keys())
    history_labels = sorted_months
    history_amounts = [round(history_map[m]['amount'], 2) for m in sorted_months]
    history_counts = [history_map[m]['count'] for m in sorted_months]

    trend_year = trend_month[:4] if trend_month and len(trend_month) >= 4 else str(now.year)
    trend_start = f"{trend_year}-01"
    trend_query = session.query(MonthlyRecord.month, func.sum(MonthlyRecord.hours)).join(Mechanic, MonthlyRecord.mechanic_id == Mechanic.id)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        trend_query = trend_query.filter(Mechanic.workshop_id == current_user.workshop_id)
    if trend_name:
        trend_query = trend_query.filter(Mechanic.name == trend_name)
    elif trend_team:
        trend_query = trend_query.filter(Mechanic.team == trend_team)
    trend_query = trend_query.filter(MonthlyRecord.month >= trend_start, MonthlyRecord.month <= trend_month).group_by(MonthlyRecord.month).order_by(MonthlyRecord.month.asc())
    trend_rows = trend_query.all()
    trend_labels = [r[0] for r in trend_rows]
    trend_values = [round(float(r[1] or 0), 2) for r in trend_rows]

    if not hours_month or hours_month == "无数据":
        hours_month = now.strftime('%Y-%m')
    hours_year = hours_month[:4] if hours_month and len(hours_month) >= 4 else str(now.year)
    hours_start = f"{hours_year}-01"
    hours_query = session.query(MonthlyRecord.month, func.sum(MonthlyRecord.hours)).join(Mechanic, MonthlyRecord.mechanic_id == Mechanic.id)
    hours_query = hours_query.filter(Mechanic.status == '在岗', Mechanic.identity == '随车机械师')
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        hours_query = hours_query.filter(Mechanic.workshop_id == current_user.workshop_id)
    if hours_mechanic:
        hours_query = hours_query.filter(Mechanic.name == hours_mechanic)
    elif hours_teams:
        hours_query = hours_query.filter(Mechanic.team.in_(hours_teams))
    hours_query = hours_query.filter(MonthlyRecord.month >= hours_start, MonthlyRecord.month <= hours_month).group_by(MonthlyRecord.month).order_by(MonthlyRecord.month.asc())
    hours_rows = hours_query.all()
    hours_trend_labels = [r[0] for r in hours_rows]
    hours_trend_values = [round(float(r[1] or 0), 2) for r in hours_rows]

    hours_mech_query = session.query(Mechanic.name, Mechanic.team).filter(Mechanic.status == '在岗', Mechanic.identity == '随车机械师')
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        hours_mech_query = hours_mech_query.filter(Mechanic.workshop_id == current_user.workshop_id)
    hours_rows_2 = hours_mech_query.all()
    hours_mechanic_list = sorted({r[0] for r in hours_rows_2 if r[0]})
    hours_team_list = sorted({r[1] for r in hours_rows_2 if r[1]})

    months_hours_query = session.query(MonthlyRecord.month).join(Mechanic, MonthlyRecord.mechanic_id == Mechanic.id)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        months_hours_query = months_hours_query.filter(Mechanic.workshop_id == current_user.workshop_id)
    months_hours = sorted({r[0] for r in months_hours_query.all() if r[0]}, reverse=True)

    def normalize_issue_month(val):
        if not val:
            return None
        s = str(val).strip()
        if not s:
            return None
        s = s.replace('年', '-').replace('月', '').replace('.', '-').replace('/', '-')
        parts = s.split('-')
        if len(parts) >= 2:
            y = parts[0]
            m = parts[1]
            if len(m) == 1:
                m = f"0{m}"
            return f"{y}-{m}"
        return None

    issues_query = session.query(Issue.date)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        issues_query = issues_query.join(Mechanic, Issue.mechanic_id == Mechanic.id).filter(Mechanic.workshop_id == current_user.workshop_id)
    issues_months_set = set()
    for d in issues_query.all():
        m = normalize_issue_month(d[0])
        if m:
            issues_months_set.add(m)
    issues_months = sorted(issues_months_set, reverse=True)

    session.close()
    return render_template(
        'index.html',
        total_mechanics=total,
        over_1000=over_1000,
        backup_count=backup_count,
        regular_count=regular_count,
        backup_mechanics=backup_count,
        regular_mechanics=regular_count,
        last_month_imported=last_month_imported,
        score_labels=json.dumps(score_labels, ensure_ascii=False),
        score_counts=json.dumps(score_counts),
        deduction_labels=json.dumps(deduction_labels, ensure_ascii=False),
        deduction_counts=json.dumps(deduction_counts),
        reward_labels=json.dumps(reward_labels, ensure_ascii=False),
        reward_counts=json.dumps(reward_counts),
        hours_labels=json.dumps(hours_labels, ensure_ascii=False),
        hours_counts=json.dumps(hours_counts),
        history_labels=json.dumps(history_labels, ensure_ascii=False),
        history_amounts=json.dumps(history_amounts),
        history_counts=json.dumps(history_counts),
        filter_year=filter_year,
        filter_start_month=filter_start_month,
        filter_end_month=filter_end_month,
        trend_month=trend_month,
        trend_name=trend_name,
        trend_team=trend_team,
        trend_labels=json.dumps(trend_labels, ensure_ascii=False),
        trend_values=json.dumps(trend_values),
        mechanic_names=mechanic_names,
        mechanic_teams=mechanic_teams,
        months_hours=months_hours,
        months_issues=issues_months,
        hours_month=hours_month,
        hours_mechanic=hours_mechanic,
        hours_teams=hours_teams,
        hours_mechanic_list=hours_mechanic_list,
        hours_team_list=hours_team_list,
        hours_trend_labels=json.dumps(hours_trend_labels, ensure_ascii=False),
        hours_trend_values=json.dumps(hours_trend_values),
        hours_months=months_hours,
        issue_months=issues_months
    )

@app.route('/logs')
@login_required
def view_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    session = Session()
    q = session.query(OperationLog).order_by(desc(OperationLog.created_at))
    
    total_count = q.count()
    logs = q.offset((page - 1) * per_page).limit(per_page).all()
    
    total_pages = (total_count + per_page - 1) // per_page
    start_index = (page - 1) * per_page + 1
    end_index = min(page * per_page, total_count)
    
    session.close()
    
    return render_template('logs.html', logs=logs,
                           page=page, per_page=per_page, total_pages=total_pages,
                           total_count=total_count, start_index=start_index, end_index=end_index)


@app.route('/upload_check', methods=['POST'])
def upload_check():
    # AJAX check for duplicates
    month = request.form.get('month')
    if not month:
        return jsonify({'status': 'error', 'message': '请选择月份'})
        
    session = Session()
    exists = session.query(MonthlyRecord).filter_by(month=month).first()
    session.close()
    
    if exists:
        return jsonify({'status': 'warning', 'message': f'{month} 月份的数据已存在！继续导入将追加/覆盖数据，是否继续？'})
    return jsonify({'status': 'ok'})

@app.route('/upload_att6', methods=['POST'])
@login_required
def upload_att6():
    if current_user.role == ROLE_SECTION:
        flash('段级权限无法导入附件6数据', 'error')
        return redirect(url_for('data_att6'))
    file = request.files.get('file')
    if not file or not file.filename:
        flash('请选择文件', 'error')
        return redirect(url_for('data_att6'))
        
    session = Session()
    try:
        path = os.path.join(UPLOAD_FOLDER, f"att6_{secure_filename(file.filename)}")
        file.save(path)
        
        # Reuse logic from init_data.py but inline here to avoid importing
        # Find header row dynamically
        try:
            # Read first few rows to find header
            # Use header=None to read raw rows
            df_preview = pd.read_excel(path, sheet_name=0, header=None, nrows=10)
            header_row_idx = -1
            
            # Look for row containing '工号'
            for i, row in df_preview.iterrows():
                # Convert row to string and check
                row_values = [str(val).strip() for val in row.values]
                if '工号' in row_values:
                    header_row_idx = i
                    break
            
            if header_row_idx == -1:
                # Fallback to default 2 if not found
                header_row_idx = 2
            
            # Read actual data with found header
            df = pd.read_excel(path, sheet_name=0, header=header_row_idx)
            
            # Normalize columns: remove newlines, strip spaces (including full-width)
            df.columns = [str(c).replace('\n', '').replace(' ', '').replace('\u3000', '').replace('\t', '').strip() for c in df.columns]
            
            print(f"DEBUG: Header found at row {header_row_idx}")
            print(f"DEBUG: Columns found: {df.columns.tolist()}")
            
            # Verify critical columns exist
            if '工号' not in df.columns:
                # Try finding fuzzy match?
                fuzzy_col = next((c for c in df.columns if '工号' in c), None)
                if fuzzy_col:
                    df.rename(columns={fuzzy_col: '工号'}, inplace=True)
                    print(f"DEBUG: Renamed column '{fuzzy_col}' to '工号'")
                else:
                    flash(f'导入失败: 未找到“工号”列。请检查Excel表头。识别到的列: {df.columns.tolist()}', 'error')
                    return redirect(url_for('data_att6'))

            # Clear existing
            session.query(RewardHistory).delete()
            session.query(Attachment6Data).delete()
            session.flush()
            session.commit()
            
            count = 0
            
            # Helpers
            def excel_date_to_str(val):
                try:
                    if pd.isna(val): return ""
                    val_str = str(val).strip()
                    if not val_str: return ""
                    try:
                        val_float = float(val)
                        if val_float > 40000 and val_float < 50000:
                            dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(val_float) - 2)
                            return dt.strftime('%Y.%m')
                        elif val_float > 202000 and val_float < 203000:
                            val_s = str(int(val_float))
                            if len(val_s) >= 6: return f"{val_s[:4]}.{val_s[4:6]}"
                    except: pass
                    if '年' in val_str and '月' in val_str:
                        val_str = val_str.replace('年', '.').replace('月', '')
                    val_str = val_str.replace('-', '.')
                    if '.' in val_str:
                        parts = val_str.split('.')
                        if len(parts) >= 2: return f"{parts[0]}.{int(parts[1]):02d}"
                    return val_str
                except: return str(val)

            def clean_str(val):
                s = str(val).strip()
                if s == 'nan' or s == '/': return ""
                return s
            
            def clean_float(val):
                try:
                    s = str(val).strip()
                    if s == '/' or s == 'nan' or not s: return 0.0
                    return float(s)
                except: return 0.0

            def clean_int(val):
                try:
                    s = str(val).strip()
                    if s == '/' or s == 'nan' or not s: return 0
                    return int(float(s))
                except: return 0
            
            for _, row in df.iterrows():
                # Use Normalized column names (no spaces, no newlines)
                if pd.isna(row.get('工号')): continue
                
                emp_id = str(row.get('工号')).strip()
                if emp_id.endswith('.0'): emp_id = emp_id[:-2]
                if not emp_id: continue

                mechanic = session.query(Mechanic).filter_by(employee_id=emp_id).first()
                if not mechanic:
                    mechanic = Mechanic(employee_id=emp_id)
                    session.add(mechanic)
                
                mechanic.name = str(row.get('姓名', '')).strip()
                mechanic.team = str(row.get('乘务组别', '')).strip()
                
                history = RewardHistory(mechanic=mechanic)
                # Map columns flexibly
                history.reward_date = excel_date_to_str(row.get('本次满足奖励时间') or row.get('本次满足奖励时间（月份）') or row.get('本次满足奖励时间(月份)'))
                # Match Amount Column
                # Possible names: "本次奖励周期奖励金额", "奖励金额", "本月奖励金额"
                amt_col_1 = next((c for c in df.columns if '本次奖励周期奖励金额' in c), None)
                if amt_col_1:
                     history.amount = clean_float(row.get(amt_col_1))
                else:
                     # Fallback to just '奖励金额'
                     amt_col_2 = next((c for c in df.columns if '奖励金额' in c and '本月' not in c), None)
                     if amt_col_2:
                         history.amount = clean_float(row.get(amt_col_2))
                     else:
                         history.amount = 0.0

                # Check deduction sign - Force Negative
                # Try variations of column name
                raw_val = clean_float(row.get('本次奖励周期内千工时累积扣分') or row.get('千工时累积扣分') or row.get('本次奖励周期内扣分'))
                final_ded = -abs(raw_val)
                # Ensure it's not -0.0
                if final_ded == 0: final_ded = 0.0
                
                history.deduction = final_ded
                
                history.cleared_hours = clean_float(row.get('累计清零工时'))
                # Fuzzy match for cleared hours
                if not history.cleared_hours:
                    cleared_col = next((c for c in df.columns if '累计清零工时' in c), None)
                    if cleared_col:
                        history.cleared_hours = clean_float(row.get(cleared_col))
                
                # 本月奖励金额 = 奖励金额 + 本次额外奖励
                # history.total_amount = clean_float(row.get(amt_col)) if amt_col else 0.0
                # User suggests: Just use DB col "本月奖励金额", AND it should equal amount + extra.
                
                amt_col = next((c for c in df.columns if '本月奖励金额' in c), None)
                if amt_col:
                    history.total_amount = clean_float(row.get(amt_col))
                else:
                    # Fallback calculation if column missing?
                    history.total_amount = history.amount + history.extra_reward
                
                con_info = row.get('连续奖励情况（上一周期奖励）') or row.get('连续奖励情况(上一周期奖励)') or row.get('连续奖励情况')
                history.consecutive_info = clean_str(con_info)
                
                session.add(history)
                
                att6 = Attachment6Data(
                    employee_id=emp_id,
                    name=mechanic.name,
                    team=mechanic.team,
                    reward_date=history.reward_date,
                    reward_cycle=history.reward_cycle,
                    reward_amount=history.amount,
                    activity_cumulative_hours=clean_float(row.get('活动起累计工时')),
                    cycle_deduction=final_ded,
                    cleared_hours=history.cleared_hours,
                    balance_hours=clean_float(row.get('当前前结余工时')),
                    past_reward_count=clean_int(row.get('过去已奖励次数')),
                    is_consecutive=clean_str(row.get('本次奖励是否触发连续奖励')),
                    consecutive_info=history.consecutive_info,
                    extra_reward=history.extra_reward,
                    total_amount=history.total_amount
                )
                session.add(att6)
                count += 1
                
                # Debug print for first few rows
                if count <= 5:
                    print(f"DEBUG: Row {count} - Raw Ded: {raw_val}, Final Ded: {final_ded}")
            
            session.commit()
            flash(f'成功导入附件6数据，共 {count} 条记录', 'success')
            
        except Exception as e:
            session.rollback()
            flash(f'导入失败: {e}', 'error')
            print(e)
            
    except Exception as e:
        flash(f'文件保存失败: {e}', 'error')
    finally:
        session.close()
        
    return redirect(url_for('data_att6'))

@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    if current_user.role == ROLE_SECTION:
        flash('段级权限无法导入数据', 'error')
        return redirect(url_for('index'))
    if request.method == 'POST':
        month = request.form.get('month')
        hours_file = request.files.get('hours_file')
        issues_file = request.files.get('issues_file')
        
        if not hours_file and not issues_file:
            flash('请至少选择一个文件上传', 'error')
            return redirect(url_for('upload'))
            
        session = Session()
        try:
            hours_added = 0
            issues_processed = 0
            deleted_issues = 0
            
            # 1. Process Hours
            if hours_file and hours_file.filename:
                h_path = os.path.join(UPLOAD_FOLDER, f"hours_{month}_{secure_filename(hours_file.filename)}")
                hours_file.save(h_path)
                
                df_hours = pd.read_excel(h_path)
                df_hours.columns = [str(c).replace('\n', ' ').strip() for c in df_hours.columns]
                
                for _, row in df_hours.iterrows():
                    emp_id = str(row.get('工号')).strip()
                    if emp_id.endswith('.0'): emp_id = emp_id[:-2]
                    
                    mechanic = session.query(Mechanic).filter_by(employee_id=emp_id).first()
                    if not mechanic:
                        mechanic = Mechanic(employee_id=emp_id)
                        mechanic.name = str(row.get('姓名', '')).strip()
                        mechanic.team = str(row.get('班组', '')).strip()
                        mechanic.total_hours = 0.0
                        mechanic.current_cycle_deduction = 0.0
                        session.add(mechanic)
                        session.flush() # Ensure ID
                    
                    # Logic: Find existing record. If exists, subtract old hours first, then add new.
                    existing_rec = session.query(MonthlyRecord).filter_by(mechanic_id=mechanic.id, month=month).first()
                    
                    new_hours = safe_float(row.get('月度工时小计'))
                    
                    if existing_rec:
                        # Subtract previous import for this month to avoid double counting
                        if mechanic.total_hours is None: mechanic.total_hours = 0.0
                        mechanic.total_hours -= existing_rec.hours
                        
                        # Update record
                        existing_rec.hours = new_hours
                        
                        # Add new hours
                        mechanic.total_hours += new_hours
                    else:
                        if mechanic.total_hours is None: mechanic.total_hours = 0.0
                        mechanic.total_hours += new_hours
                        rec = MonthlyRecord(mechanic=mechanic, month=month, hours=new_hours)
                        session.add(rec)
                    
                    hours_added += 1
            
            # 2. Process Issues
            if issues_file and issues_file.filename:
                i_path = os.path.join(UPLOAD_FOLDER, f"issues_{month}_{secure_filename(issues_file.filename)}")
                issues_file.save(i_path)
                
                df_issues = pd.read_excel(i_path)
                df_issues.columns = [str(c).replace('\n', ' ').strip() for c in df_issues.columns]
                
                # Overwrite Logic:
                # User says: "If not deleted... overwrite".
                # To ensure overwrite, we must delete ALL issues for this month first.
                # However, Issue table has `date` (YYYY-MM-DD or numeric), not explicit `month`.
                # We should delete issues where date matches `month` (YYYY-MM).
                
                # Identify issues to delete for this month
                # Reusing is_in_month helper logic (inline here for simplicity)
                def is_in_month_check(date_str, target_month):
                    if not date_str: return False
                    try:
                        d_s = str(date_str)
                        if d_s.startswith(target_month): return True
                        if d_s.replace('.','').isdigit() and float(d_s) > 30000:
                            dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(float(d_s)) - 2)
                            return dt.strftime('%Y-%m') == target_month
                        target_dot = target_month.replace('-', '.')
                        if d_s.startswith(target_dot): return True
                        target_slash = target_month.replace('-', '/')
                        if d_s.startswith(target_slash): return True
                        return False
                    except: return False

                # Find all issues matching this month and delete them
                all_existing_issues = session.query(Issue).all()
                for issue in all_existing_issues:
                    if is_in_month_check(issue.date, month):
                        # Revert mechanic deduction before deleting
                        if issue.mechanic:
                            issue.mechanic.current_cycle_deduction -= abs(issue.detail or 0)
                            if issue.mechanic.current_cycle_deduction < 0: issue.mechanic.current_cycle_deduction = 0
                        session.delete(issue)
                        deleted_issues += 1
                
                # Flush to ensure deletions are processed before adding new ones
                session.flush()
                
                for _, row in df_issues.iterrows():
                    name = str(row.get('姓名', '')).strip()
                    mechanics = session.query(Mechanic).filter_by(name=name).all()
                    if not mechanics: continue
                    mechanic = mechanics[0]
                    
                    deduction = safe_float(row.get('扣分明细')) # e.g. -0.5
                    deduction_mag = abs(deduction)
                    problem = str(row.get('问题', ''))
                    date_str = str(row.get('检查日期', ''))
                    
                    # Check duplicate issue (Now technically redundant if we cleared month, but safe to keep for safety within same file)
                    dup = session.query(Issue).filter_by(
                        mechanic_id=mechanic.id, 
                        problem=problem, 
                        detail=deduction
                    ).first()
                    
                    if dup:
                        continue
                    
                    # Create Issue
                    issue = Issue(
                        mechanic_id=mechanic.id,
                        date=date_str,
                        problem=problem,
                        source=str(row.get('问题来源', '')),
                        clause=str(row.get('扣分条款', '')),
                        detail=deduction,
                        total_deduction=deduction,
                        status='未结算'
                    )
                    session.add(issue)
                    
                    if mechanic.current_cycle_deduction is None: mechanic.current_cycle_deduction = 0.0
                    mechanic.current_cycle_deduction += deduction_mag
                    
                    # Update MonthlyRecord (Create if not exists)
                    rec = session.query(MonthlyRecord).filter_by(mechanic_id=mechanic.id, month=month).first()
                    if not rec:
                        rec = MonthlyRecord(mechanic=mechanic, month=month)
                        session.add(rec)
                    
                    if rec.deduction is None: rec.deduction = 0.0
                    rec.deduction += deduction_mag
                    rec.issues_details = (rec.issues_details or "") + f"{problem} ({deduction}); "
                    issues_processed += 1
                
            write_operation_log(
                "导入月度数据",
                json.dumps({
                    "month": month,
                    "hours_added": hours_added,
                    "issues_added": issues_processed,
                    "issues_deleted": deleted_issues,
                    "hours_file": hours_file.filename if hours_file else "",
                    "issues_file": issues_file.filename if issues_file else ""
                }, ensure_ascii=False),
                session=session
            )
            session.commit()
            
            msg = []
            if hours_added > 0: msg.append(f'{hours_added} 条工时记录')
            if issues_processed > 0: msg.append(f'{issues_processed} 条新问题记录')
            
            if msg:
                flash(f'成功处理: {", ".join(msg)}', 'success')
            else:
                flash('未处理任何数据', 'warning')
            
        except Exception as e:
            session.rollback()
            flash(f'处理失败: {str(e)}', 'error')
            print(e)
        finally:
            session.close()
            
        return redirect(url_for('index'))
        
    return render_template('upload.html')

@app.route('/mechanics')
@login_required
def mechanics():
    sort_by = request.args.get('sort', 'employee_id')
    order = request.args.get('order', 'asc')
    search_query = request.args.get('search', '')
    team_filter = [t for t in request.args.getlist('team') if t]
    include_backup = request.args.get('include_backup', '') == '1'
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    settle_progress = flask_session.pop('settle_progress', None)
    
    session = Session()
    q = session.query(Mechanic).options(joinedload(Mechanic.rewards_history), joinedload(Mechanic.issues), joinedload(Mechanic.workshop))
    
    # Workshop Filter
    if current_user.role == ROLE_WORKSHOP:
        if current_user.workshop_id:
            q = q.filter(Mechanic.workshop_id == current_user.workshop_id)

    team_query = session.query(Mechanic.team)
    if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        team_query = team_query.filter(Mechanic.workshop_id == current_user.workshop_id)
    team_list = sorted({t[0] for t in team_query.all() if t[0]})
            
    if search_query:
        q = q.filter(or_(Mechanic.name.contains(search_query), Mechanic.employee_id.contains(search_query)))
    if team_filter:
        q = q.filter(Mechanic.team.in_(team_filter))
    if not include_backup and '后备' not in (team_filter or []):
        q = q.filter(or_(Mechanic.team != '后备', Mechanic.team.is_(None)))
    
    if order == 'asc':
        q = q.order_by(asc(getattr(Mechanic, sort_by)))
    else:
        q = q.order_by(desc(getattr(Mechanic, sort_by)))
    
    # Check if coming from dashboard >1000 link
    filter_type = request.args.get('filter')
    if filter_type == 'over_1000':
        # Sort by hours desc, but user might have clicked header sort.
        # If user explicitly sorted, respect that. If not (default sort), force hours desc.
        if sort_by == 'employee_id' and order == 'asc': # Default
            q = q.order_by(desc(Mechanic.total_hours))
            sort_by = 'total_hours'
            order = 'desc'
    
    # Pagination
    total_count = q.count()
    mechanics = q.offset((page - 1) * per_page).limit(per_page).all()
    
    total_pages = (total_count + per_page - 1) // per_page
    start_index = (page - 1) * per_page + 1
    end_index = min(page * per_page, total_count)
    
    # Logic for "Imported Hours (Month)" column
    # 1. Find latest imported month
    last_record = session.query(MonthlyRecord).order_by(MonthlyRecord.month.desc()).first()
    last_import_month = last_record.month if last_record else "无数据"
    
    # Calculate Previous Month
    prev_import_month = "无数据"
    if last_record:
        try:
            # Assuming format YYYY-MM
            y, m = map(int, last_import_month.split('-'))
            if m == 1:
                prev_y = y - 1
                prev_m = 12
            else:
                prev_y = y
                prev_m = m - 1
            prev_import_month = f"{prev_y}-{prev_m:02d}"
        except:
            pass

    # 2. Get hours for this month and previous month
    import_map = {}
    prev_import_map = {}
    
    if last_record:
        # Current Month
        recs = session.query(MonthlyRecord).filter_by(month=last_import_month).all()
        for r in recs:
            import_map[r.mechanic_id] = r.hours
            
        # Previous Month
        recs_prev = session.query(MonthlyRecord).filter_by(month=prev_import_month).all()
        for r in recs_prev:
            prev_import_map[r.mechanic_id] = r.hours
    
    activity_map = {}
    if mechanics:
        employee_ids = [m.employee_id for m in mechanics if m.employee_id]
        base_map = {}
        if employee_ids:
            rows_att6 = session.query(Attachment6Data).filter(
                Attachment6Data.employee_id.in_(employee_ids)
            ).order_by(Attachment6Data.employee_id, Attachment6Data.id.desc()).all()
            for r in rows_att6:
                if r.employee_id not in base_map:
                    base_map[r.employee_id] = (r.reward_date, r.activity_cumulative_hours or 0.0)
        for m in mechanics:
            base_val = m.base_hours if m.base_hours is not None else 0.0
            last_reward_month = None
            base_info = base_map.get(m.employee_id)
            if base_info:
                last_reward_month = base_info[0]
                base_val = base_info[1]
            qh = session.query(func.sum(MonthlyRecord.hours)).filter(
                MonthlyRecord.mechanic_id == m.id
            )
            if last_reward_month:
                qh = qh.filter(MonthlyRecord.month > last_reward_month)
            if last_record:
                qh = qh.filter(MonthlyRecord.month <= last_import_month)
            extra = qh.scalar() or 0.0
            activity_map[m.id] = base_val + extra
    
    session.close()
    return render_template('mechanics.html', mechanics=mechanics, 
                           sort_by=sort_by, order=order, search_query=search_query,
                           team_list=team_list, team_filter=team_filter, include_backup=include_backup,
                           page=page, per_page=per_page, total_pages=total_pages,
                           total_count=total_count, start_index=start_index, end_index=end_index,
                           last_import_month=last_import_month, import_map=import_map,
                           prev_import_month=prev_import_month, prev_import_map=prev_import_map,
                           activity_map=activity_map, settle_progress=settle_progress)

@app.route('/download_mechanic_template')
@login_required
def download_mechanic_template():
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "职工导入模板"
    headers = ['工号', '姓名', '车间', '班组', '身份', '状态']
    ws.append(headers)
    ws.append(['12345', '张三', '上海南动车所', '一班', '随车机械师', '在岗'])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='职工导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/import_mechanics', methods=['POST'])
@login_required
def import_mechanics():
    from openpyxl import load_workbook
    if current_user.role == ROLE_SECTION:
        flash('段级权限无法导入职工数据', 'error')
        return redirect(url_for('mechanics'))
        
    if 'file' not in request.files:
        flash('未上传文件', 'error')
        return redirect(url_for('mechanics'))
        
    file = request.files['file']
    if file.filename == '':
        flash('未选择文件', 'error')
        return redirect(url_for('mechanics'))
        
    if file and file.filename.endswith('.xlsx'):
        try:
            wb = load_workbook(file)
            ws = wb.active
            rows = list(ws.rows)
            
            if not rows:
                flash('文件为空', 'error')
                return redirect(url_for('mechanics'))
                
            header = [cell.value for cell in rows[0]]
            required = ['工号', '姓名', '车间', '班组', '身份', '状态']
            mapping = {}
            for i, h in enumerate(header):
                if h in required:
                    mapping[h] = i
            
            if len(mapping) < len(required):
                flash(f'模板格式错误，缺少列: {set(required) - set(mapping.keys())}', 'error')
                return redirect(url_for('mechanics'))
            
            session = Session()
            added_count = 0
            updated_count = 0
            
            workshops = {w.name: w for w in session.query(Workshop).all()}
            
            for row in rows[1:]:
                emp_id = str(row[mapping['工号']].value).strip() if row[mapping['工号']].value else None
                name = str(row[mapping['姓名']].value).strip() if row[mapping['姓名']].value else None
                workshop_name = str(row[mapping['车间']].value).strip() if row[mapping['车间']].value else None
                team = str(row[mapping['班组']].value).strip() if row[mapping['班组']].value else None
                identity = str(row[mapping['身份']].value).strip() if row[mapping['身份']].value else None
                status = str(row[mapping['状态']].value).strip() if row[mapping['状态']].value else None
                
                if not emp_id or not name:
                    continue
                    
                target_workshop = None
                if workshop_name:
                    if workshop_name in workshops:
                        target_workshop = workshops[workshop_name]
                    else:
                        new_w = Workshop(name=workshop_name)
                        session.add(new_w)
                        session.flush()
                        workshops[workshop_name] = new_w
                        target_workshop = new_w
                
                if current_user.role == ROLE_WORKSHOP:
                    if target_workshop and target_workshop.id != current_user.workshop_id:
                        continue
                    if not target_workshop:
                        target_workshop = session.query(Workshop).get(current_user.workshop_id)

                mech = session.query(Mechanic).filter_by(employee_id=emp_id).first()
                if mech:
                    mech.name = name
                    mech.team = team
                    mech.identity = identity
                    mech.status = status
                    if target_workshop:
                        mech.workshop_id = target_workshop.id
                    updated_count += 1
                else:
                    mech = Mechanic(
                        employee_id=emp_id,
                        name=name,
                        team=team,
                        identity=identity,
                        status=status,
                        workshop_id=target_workshop.id if target_workshop else None
                    )
                    session.add(mech)
                    added_count += 1
            
            session.commit()
            session.close()
            flash(f'导入完成: 新增 {added_count} 人, 更新 {updated_count} 人', 'success')
            
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'error')
            
    return redirect(url_for('mechanics'))

@app.route('/export/cycle')
def export_cycle_preview():
    # Logic for Attachment 2-1 (Cycle Stats)
    # This seems to be a summary of ALL mechanics for the current cycle?
    # Or is it individual?
    # Template 2-1: 序号, 姓名, 工号, 乘务组别, 当前积分, 问题描述, 问题来源, 扣分条款, 扣分明细, 扣分总计
    # "当前积分" -> Current Cycle Points? Or Base (1000/12) - Deduction?
    # Assuming "积分" usually starts at 12 or 1000 per cycle.
    # Let's assume standard 12 points for safety or 1000 for competition.
    # Since it's "保安全竞赛", likely 1000 base.
    
    session = Session()
    mechanics = session.query(Mechanic).filter(Mechanic.current_cycle_deduction > 0).all()
    # Or everyone? Usually stats include everyone.
    # If "Cycle Stats", probably everyone with deductions or just everyone.
    # Let's show everyone who has issues first, or everyone.
    # User said "按照模板进行显示".
    # Template has "问题描述...". If no issues, these are blank.
    
    # Let's fetch everyone for now, but paginate or just show all?
    # Export usually implies all.
    mechanics_q = session.query(Mechanic).options(joinedload(Mechanic.issues))
    if getattr(current_user, 'is_authenticated', False) and current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        mechanics_q = mechanics_q.filter(Mechanic.workshop_id == current_user.workshop_id)
    mechanics = mechanics_q.all()
    session.close()
    return render_template('stats_cycle.html', mechanics=mechanics)

@app.route('/export/cycle/download')
def export_cycle_download():
    session = Session()
    mechanics_q = session.query(Mechanic).options(joinedload(Mechanic.issues), joinedload(Mechanic.rewards_history))
    if getattr(current_user, 'is_authenticated', False) and current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        mechanics_q = mechanics_q.filter(Mechanic.workshop_id == current_user.workshop_id)
    mechanics = mechanics_q.all()
    
    template_path = os.path.join('excel_templates', '附件2-2模板.xlsx')
    output_path = os.path.join('excel_templates', 'generated_cycle_stats.xlsx')
    shutil.copy(template_path, output_path)
    
    import openpyxl
    import re
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    last_record = session.query(MonthlyRecord).order_by(MonthlyRecord.month.desc()).first()
    settle_month = last_record.month if last_record else datetime.now().strftime('%Y-%m')
    center_align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    left_align = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    left_wrap = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
    base_font = openpyxl.styles.Font(size=11)
    thin = openpyxl.styles.Side(style='thin')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    
    row_idx = 4 # Data starts row 4 based on template (Index 3 is Row 4)
    # Check if Row 4 is merged? If so, unmerge it or check template.
    # User error "MergedCell object attribute value is read-only" implies we are trying to write to a merged cell.
    # The template might have pre-filled merged cells for data rows?
    # Or maybe the header (Row 1-3) has merged cells and we are accidentally writing to them?
    # We start at row_idx=4.
    # Let's inspect if the template has merged cells in the data area.
    # If so, we should probably unmerge them or ensure we write to the top-left cell.
    # But usually data rows shouldn't be merged in a template unless it's just one example row.
    # Safest way: Unmerge data area before writing?
    # Or check if cell is MergedCell.
    
    # If the template has existing data rows (e.g. example row) that are merged, we might be hitting that.
    # Let's try to unmerge the target range for the current row before writing?
    # Or just handle MergedCell.
    
    # Unmerge any potential merged cells in data area
    # Check intersection with data rows (row_idx onwards)
    for range_ in list(ws.merged_cells.ranges):
        if range_.max_row >= row_idx:
            ws.unmerge_cells(str(range_))
            
    def extract_ym(value):
        if not value:
            return None
        m = re.search(r'(\d{4})\D*(\d{1,2})', str(value))
        if not m:
            return None
        return f"{int(m.group(1))}年{int(m.group(2)):02d}月"

    def build_cycle(last_cycle):
        start_part = last_cycle
        if last_cycle and '-' in last_cycle:
            start_part = last_cycle.split('-')[0]
        start_fmt = extract_ym(start_part) or "2025年12月"
        end_fmt = extract_ym(settle_month) or extract_ym(datetime.now().strftime('%Y-%m')) or str(settle_month)
        return f"{start_fmt}-{end_fmt}"

    ws.column_dimensions['G'].width = 80

    i = 1
    for m in mechanics:
        issues = [issue for issue in m.issues if getattr(issue, 'status', '未结算') == '未结算']
        cycle_str = build_cycle(m.rewards_history[0].reward_cycle if m.rewards_history else None)
        if not issues:
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=m.name)
            ws.cell(row=row_idx, column=3, value=m.employee_id)
            ws.cell(row=row_idx, column=4, value=m.team)
            ws.cell(row=row_idx, column=5, value=cycle_str)
            ws.cell(row=row_idx, column=6, value=m.total_hours)
            ws.cell(row=row_idx, column=11, value=0)
            for col in range(1, 12):
                cell = ws.cell(row=row_idx, column=col)
                if col == 7:
                    cell.alignment = left_wrap
                elif 7 <= col <= 10:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.font = base_font
                cell.border = border
            row_idx += 1
            i += 1
            continue
        start_row = row_idx
        total_ded = -sum(abs(issue.detail or 0) for issue in issues)
        for issue in issues:
            detail_val = -abs(issue.detail or 0)
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=m.name)
            ws.cell(row=row_idx, column=3, value=m.employee_id)
            ws.cell(row=row_idx, column=4, value=m.team)
            ws.cell(row=row_idx, column=5, value=cycle_str)
            ws.cell(row=row_idx, column=6, value=m.total_hours)
            ws.cell(row=row_idx, column=7, value=issue.problem or "")
            ws.cell(row=row_idx, column=8, value=issue.source or "")
            ws.cell(row=row_idx, column=9, value=issue.clause or "")
            ws.cell(row=row_idx, column=10, value=detail_val)
            for col in range(1, 12):
                cell = ws.cell(row=row_idx, column=col)
                if col == 7:
                    cell.alignment = left_wrap
                elif 7 <= col <= 10:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.font = base_font
                cell.border = border
            row_idx += 1
        end_row = row_idx - 1
        ws.cell(row=start_row, column=11, value=total_ded)
        for col in range(1, 7):
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            ws.cell(row=start_row, column=col).alignment = center_align
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
        ws.cell(row=start_row, column=11).alignment = center_align
        for r in range(start_row, end_row + 1):
            for c in range(1, 12):
                cell = ws.cell(row=r, column=c)
                if c == 7:
                    cell.alignment = left_wrap
                elif 7 <= c <= 10:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.font = base_font
                cell.border = border
        i += 1
        
    wb.save(output_path)
    session.close()
    return send_file(output_path, as_attachment=True, download_name="周期积分统计.xlsx")

@app.route('/export/annual')
def export_annual_preview():
    # Logic for Attachment 2-2 (Annual/Thousand Hour Stats)
    # Template 2-2: 序号, 姓名, 工号, 乘务组别, 统计周期, 累计工时, 竞赛周期内扣分明细...
    # Similar to Mechanic List but with issues.
    session = Session()
    mechanics_q = session.query(Mechanic).options(joinedload(Mechanic.issues))
    if getattr(current_user, 'is_authenticated', False) and current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        mechanics_q = mechanics_q.filter(Mechanic.workshop_id == current_user.workshop_id)
    mechanics = mechanics_q.all()
    session.close()
    return render_template('stats_annual.html', mechanics=mechanics)

@app.route('/export/annual/download')
def export_annual_download():
    session = Session()
    mechanics_q = session.query(Mechanic).options(joinedload(Mechanic.issues), joinedload(Mechanic.rewards_history))
    if getattr(current_user, 'is_authenticated', False) and current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        mechanics_q = mechanics_q.filter(Mechanic.workshop_id == current_user.workshop_id)
    mechanics = mechanics_q.all()
    
    template_path = os.path.join('excel_templates', '附件2-1模板.xlsx')
    output_path = os.path.join('excel_templates', 'generated_annual_stats.xlsx')
    shutil.copy(template_path, output_path)
    
    import openpyxl
    import re
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    center_align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    left_align = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    left_wrap = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
    base_font = openpyxl.styles.Font(size=11)
    thin = openpyxl.styles.Side(style='thin')
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    
    row_idx = 4 # Start row 4?
    # 2-2 check:
    # 0: Title
    # 1: Header
    # 2: NaN
    # Data likely starts row 3 (Index 2) or 4?
    # Let's assume Row 3 if header is Row 2.
    # Previous check showed Header at Index 1 (Row 2).
    # So Data starts Row 3.
    row_idx = 4
    
    for range_ in list(ws.merged_cells.ranges):
        if range_.max_row >= row_idx:
            ws.unmerge_cells(str(range_))

    ws.column_dimensions['F'].width = 80
            
    i = 1
    for m in mechanics:
        # Annual export should filter by include_in_annual flag, regardless of settlement status?
        # User said: "remove last year's problems... in cycle points they cannot be deleted... so give a tag 'include in annual'".
        # "Yes means new year... No means past year... Yes shows in annual export... No does not show."
        # "Does not affect cycle points unless settled."
        # So here we filter ONLY by include_in_annual.
        # But wait, typically settled issues ARE part of annual score (you lost points).
        # Unsettled issues are ALSO part of annual score (you will lose points).
        # So we should include ALL issues where include_in_annual is True.
        # However, the previous code filtered by 'status' == '未结算'.
        # If I change this to include settled issues, the 'current_score' calculation will change.
        # Usually annual score = 12 - sum(all deductions in year).
        # If the system was only counting unsettled issues, that might have been a "current pending state".
        # But for an "Annual Report", it usually lists all deductions.
        # Let's assume the user wants "include_in_annual" to be the primary filter.
        # And usually, "Annual Points" implies the final score or current running score.
        # If I include settled issues, I should sum them up.
        # But if the previous logic only showed unsettled, maybe the "settled" ones are removed from the list?
        # Re-reading user request: "Cycle points cannot delete... Annual points calculated once a year... Past year problems should be eliminated".
        # This implies that "Cycle Points" accumulates everything until settled.
        # "Annual Points" accumulates everything for the current year (settled or not).
        # So I will filter by `include_in_annual == True`.
        
        issues = [issue for issue in m.issues if getattr(issue, 'include_in_annual', True)]
        
        # Calculate score based on these issues
        total_ded_magnitude = sum(abs(issue.detail or 0) for issue in issues)
        current_score = 12 - total_ded_magnitude
        
        if not issues:
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=m.name)
            ws.cell(row=row_idx, column=3, value=m.employee_id)
            ws.cell(row=row_idx, column=4, value=m.team)
            ws.cell(row=row_idx, column=5, value=current_score)
            ws.cell(row=row_idx, column=10, value=0)
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                if col == 6:
                    cell.alignment = left_wrap
                elif col == 8:
                    cell.alignment = center_align
                elif col in (7, 9):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.font = base_font
                cell.border = border
            row_idx += 1
            i += 1
            continue
        start_row = row_idx
        total_ded = -sum(abs(issue.detail or 0) for issue in issues)
        for issue in issues:
            detail_val = -abs(issue.detail or 0)
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=m.name)
            ws.cell(row=row_idx, column=3, value=m.employee_id)
            ws.cell(row=row_idx, column=4, value=m.team)
            ws.cell(row=row_idx, column=5, value=current_score)
            ws.cell(row=row_idx, column=6, value=issue.problem or "")
            ws.cell(row=row_idx, column=7, value=issue.source or "")
            ws.cell(row=row_idx, column=8, value=issue.clause or "")
            ws.cell(row=row_idx, column=9, value=detail_val)
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                if col == 6:
                    cell.alignment = left_wrap
                elif col == 8:
                    cell.alignment = center_align
                elif col in (7, 9):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.font = base_font
                cell.border = border
            row_idx += 1
        end_row = row_idx - 1
        ws.cell(row=start_row, column=10, value=total_ded)
        for col in range(1, 6):
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            ws.cell(row=start_row, column=col).alignment = center_align
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
        ws.cell(row=start_row, column=10).alignment = center_align
        for r in range(start_row, end_row + 1):
            for c in range(1, 11):
                cell = ws.cell(row=r, column=c)
                if c == 6:
                    cell.alignment = left_wrap
                elif c == 8:
                    cell.alignment = center_align
                elif c in (7, 9):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.font = base_font
                cell.border = border
        i += 1
        
    wb.save(output_path)
    session.close()
    return send_file(output_path, as_attachment=True, download_name="年度积分统计.xlsx")

@app.route('/settle_rewards', methods=['POST'])
def settle_rewards():
    session = Session()
    try:
        # Find mechanics > 1000 hours
        # Note: Should we only settle for "随车机械师" + "在岗"?
        # User said "If > 1000, subtract 1000, add history".
        # Let's apply filters to be safe, or just hours?
        # Usually rewards are only for qualified mechanics.
        
        query = session.query(Mechanic).filter(
            Mechanic.total_hours >= 1000,
            Mechanic.identity == '随车机械师',
            Mechanic.status == '在岗'
        )
        
        # Workshop Filter
        if current_user.role == ROLE_WORKSHOP:
            if current_user.workshop_id:
                query = query.filter(Mechanic.workshop_id == current_user.workshop_id)
                
        # Sort candidates to ensure processing order (though DB insertion ID order depends on execution)
        query = query.order_by(Mechanic.name.asc())
        
        candidates = query.all()
        
        count = 0
        
        last_record = session.query(MonthlyRecord).order_by(MonthlyRecord.month.desc()).first()
        current_month = last_record.month if last_record else datetime.now().strftime('%Y-%m')
        
        candidate_ids = [m.id for m in candidates]
        monthly_hours_map = {}
        if candidate_ids:
            rows = session.query(MonthlyRecord.mechanic_id, func.sum(MonthlyRecord.hours)).filter(
                MonthlyRecord.mechanic_id.in_(candidate_ids),
                MonthlyRecord.month <= current_month
            ).group_by(MonthlyRecord.mechanic_id).all()
            for mid, total in rows:
                monthly_hours_map[mid] = total or 0.0
        
        issue_deleted_count = 0
        issue_mech_count = 0
        settle_details = []

        def get_next_cycle_for_settle(m, target_month_str):
            try:
                target_date = datetime.strptime(target_month_str, '%Y-%m')
                target_fmt = f"{target_date.year}年{target_date.month:02d}月"
                start_fmt = "2025年12月"
                if m.rewards_history:
                    last_cycle = m.rewards_history[0].reward_cycle
                    if '-' in last_cycle:
                        parts_dash = last_cycle.split('-')
                        end_part = parts_dash[-1].strip()
                        calc_dt = None
                        clean_end = end_part.replace('年', '.').replace('月', '').strip()
                        try:
                            calc_dt = datetime.strptime(clean_end, '%Y.%m')
                        except:
                            try:
                                calc_dt = datetime.strptime(clean_end, '%Y-%m')
                            except:
                                try:
                                    if '.' in clean_end:
                                        ps = clean_end.split('.')
                                        calc_dt = datetime(int(ps[0]), int(ps[1]), 1)
                                    elif '-' in clean_end:
                                        ps = clean_end.split('-')
                                        calc_dt = datetime(int(ps[0]), int(ps[1]), 1)
                                except:
                                    calc_dt = None
                        if calc_dt:
                            if calc_dt.month == 12:
                                next_dt = datetime(calc_dt.year + 1, 1, 1)
                            else:
                                next_dt = datetime(calc_dt.year, calc_dt.month + 1, 1)
                            start_fmt = f"{next_dt.year}年{next_dt.month:02d}月"
                return f"{start_fmt}-{target_fmt}"
            except:
                return target_month_str
        
        for m in candidates:
            hist_count_before = len(m.rewards_history) if m.rewards_history else 0
            # 1. Subtract Hours
            hours_before = m.total_hours or 0.0
            m.total_hours -= 1000
            hours_after = m.total_hours
            
            # Calculate deduction from issues dynamically
            # Sum of details is likely negative (e.g. -2).
            # We need magnitude for logic (e.g. 2).
            pending_issues = [i for i in m.issues if getattr(i, 'status', '未结算') == '未结算']
            total_deduction_val = sum(i.detail for i in pending_issues) if pending_issues else 0.0
            deduction_magnitude = abs(total_deduction_val)
            
            # 2. Add History
            reward_amt = 0
            if deduction_magnitude == 0:
                reward_amt = 1200
            elif deduction_magnitude <= 1:
                reward_amt = 1000
            elif deduction_magnitude < 3:
                reward_amt = 600
            elif deduction_magnitude < 6:
                reward_amt = 200
            else:
                reward_amt = 0
                
            extra_amt = 0
            is_consecutive = "否"
            last_reward = None
            if m.rewards_history:
                last_reward = m.rewards_history[0]
                
            if last_reward:
                cur_ded_val = -abs(deduction_magnitude)
                last_ded_val = -abs(last_reward.deduction or 0.0)
                last_triggered = (last_reward.extra_reward or 0) > 0
                if cur_ded_val <= -1 or last_ded_val <= -1 or last_triggered:
                    is_consecutive = "否"
                    extra_amt = 0
                else:
                    is_consecutive = "是"
                    if cur_ded_val == 0 and last_ded_val == 0:
                        extra_amt = 2000
                    else:
                        extra_amt = 1500
            
            cycle_str = get_next_cycle_for_settle(m, current_month)

            history = RewardHistory(
                mechanic_id=m.id,
                reward_date=current_month,
                reward_cycle=cycle_str,
                deduction=-deduction_magnitude if deduction_magnitude > 0 else 0.0,
                amount=reward_amt,
                cleared_hours=m.cleared_hours,
                extra_reward=extra_amt,
                total_amount=reward_amt + extra_amt,
                consecutive_info=str(last_reward.extra_reward) if last_reward else "0.0"
            )
            session.add(history)
            
            base_hours = m.base_hours if m.base_hours is not None else 0.0
            month_hours = monthly_hours_map.get(m.id, 0.0)
            act_cumulative = base_hours + month_hours
            att6 = Attachment6Data(
                employee_id=m.employee_id,
                name=m.name,
                team=m.team,
                reward_date=current_month,
                reward_cycle=cycle_str,
                reward_amount=reward_amt,
                activity_cumulative_hours=round(act_cumulative, 2),
                cycle_deduction=history.deduction,
                cleared_hours=m.cleared_hours if m.cleared_hours is not None else 0.0,
                balance_hours=m.total_hours if m.total_hours is not None else 0.0,
                past_reward_count=hist_count_before,
                is_consecutive=is_consecutive,
                consecutive_info=history.consecutive_info,
                extra_reward=extra_amt,
                total_amount=reward_amt + extra_amt
            )
            session.add(att6)
            
            issue_count_for_m = len(pending_issues)
            if issue_count_for_m > 0:
                issue_mech_count += 1
                issue_deleted_count += issue_count_for_m
            for issue in pending_issues:
                issue.status = '已结算'
            
            m.current_cycle_deduction = 0.0
            
            settle_details.append({
                "mechanic_id": m.id,
                "employee_id": m.employee_id,
                "name": m.name,
                "hours_before": hours_before,
                "hours_after": hours_after,
                "deduction": -deduction_magnitude if deduction_magnitude > 0 else 0.0,
                "reward_amount": reward_amt,
                "extra_reward": extra_amt,
                "total_amount": reward_amt + extra_amt,
                "issue_count": issue_count_for_m
            })
            count += 1
        
        flask_session['settle_progress'] = {
            'month': current_month,
            'candidate_count': count,
            'att6_count': count,
            'history_updated_count': count,
            'issue_mech_count': issue_mech_count,
            'issue_deleted_count': issue_deleted_count
        }
        
        write_operation_log(
            "本月结算",
            json.dumps({
                "month": current_month,
                "count": count,
                "issue_mech_count": issue_mech_count,
                "issue_deleted_count": issue_deleted_count,
                "details": settle_details
            }, ensure_ascii=False),
            session=session
        )
        session.commit()
        flash(f'成功结算 {count} 人，本次结算月份 {current_month}', 'success')
    except Exception as e:
        session.rollback()
        flash(f'结算失败: {e}', 'error')
    finally:
        session.close()
        
    return redirect(url_for('mechanics'))


@app.route('/rollback_settlement', methods=['POST'])
def rollback_settlement():
    # 1. Identify Target Month (Latest Imported Month)
    session = Session()
    checklist = []
    
    try:
        last_record = session.query(MonthlyRecord).order_by(MonthlyRecord.month.desc()).first()
        if not last_record:
            flash("没有可回退的数据", "warning")
            session.close()
            return redirect(url_for('mechanics'))
            
        target_month = last_record.month
        checklist.append(f"目标月份: {target_month}")
        
        # 2. Delete Monthly Records for Target Month
        # This removes the hours from the database
        count = session.query(MonthlyRecord).filter_by(month=target_month).delete()
        session.commit()
        checklist.append(f"✅ 清除工时记录: 已删除 {count} 条")
        
        # 3. Archive Uploaded Files for Target Month
        # Move files to backup folder so they aren't re-imported
        import shutil
        backup_dir = os.path.join(UPLOAD_FOLDER, 'backup', target_month)
        os.makedirs(backup_dir, exist_ok=True)
        
        moved_count = 0
        for f in os.listdir(UPLOAD_FOLDER):
            # Match YYYY-MM in filename (e.g. hours_2026-01.xlsx)
            if target_month in f and f.endswith('.xlsx'):
                src = os.path.join(UPLOAD_FOLDER, f)
                dst = os.path.join(backup_dir, f)
                try:
                    shutil.move(src, dst)
                    moved_count += 1
                except Exception as e:
                    checklist.append(f"⚠️ 文件归档失败 ({f}): {e}")
        
        if moved_count > 0:
            checklist.append(f"✅ 归档上传文件: 已移走 {moved_count} 个文件")
        else:
            checklist.append("ℹ️ 未发现需归档的文件 (跳过)")
            
        session.close() # Close session before subprocess
        
        # 4. Run init_data.py
        # This resets base hours and recalculates total_hours from remaining MonthlyRecords
        import subprocess
        import sys
        
        try:
            subprocess.check_call([sys.executable, 'init_data.py'])
            checklist.append("✅ 重置基础数据: 完成 (Total Hours已重新计算)")
        except Exception as e:
            checklist.append(f"❌ 重置基础数据失败: {e}")
            raise e
            
        # 5. Run reimport_uploads.py
        # This re-imports issues from remaining files in uploads/
        try:
            subprocess.check_call([sys.executable, 'reimport_uploads.py'])
            checklist.append("✅ 重新导入历史问题: 完成")
        except Exception as e:
            checklist.append(f"❌ 重新导入问题失败: {e}")
            raise e

        # 6. Run reload_att6.py
        # Restores reward history from attachment 6 file
        try:
            subprocess.check_call([sys.executable, 'reload_att6.py'])
            checklist.append("✅ 恢复奖励历史: 完成")
        except Exception as e:
            checklist.append(f"❌ 恢复奖励历史失败: {e}")
            raise e
            
        # Format checklist for display
        flash_html = "<br>".join(checklist)
        flash(f'数据回退成功!<br><div class="text-sm mt-2">{flash_html}</div>', 'success')
        
    except Exception as e:
        session.rollback()
        flash(f'数据回退失败: {e}', 'error')
        if 'checklist' in locals() and checklist:
             flash(f'已完成步骤: {", ".join(checklist)}', 'info')
    finally:
        # Session already closed above, but safe to ensure
        pass

    return redirect(url_for('mechanics'))

@app.route('/admin/mechanics', methods=['GET', 'POST'])
@login_required
def admin_mechanics():
    session = Session()
    search = request.args.get('search', '').strip()
    query = session.query(Mechanic).options(joinedload(Mechanic.workshop))
    
    # Workshop Filter
    if current_user.role == ROLE_WORKSHOP:
        if current_user.workshop_id:
            query = query.filter(Mechanic.workshop_id == current_user.workshop_id)
            
    if search:
        like = f"%{search}%"
        query = query.filter(or_(Mechanic.name.like(like), Mechanic.employee_id.like(like)))
    mechanics = query.all()
    session.close()
    return render_template('admin_mechanics.html', mechanics=mechanics, search_query=search)

@app.route('/admin/mechanics/edit/<int:mid>', methods=['GET', 'POST'])
@login_required
def edit_mechanic(mid):
    session = Session()
    m = session.query(Mechanic).get(mid)
    if not m:
        session.close()
        flash('人员不存在', 'error')
        return redirect(url_for('admin_mechanics'))
    if request.method == 'POST':
        m.name = request.form.get('name')
        m.employee_id = request.form.get('employee_id')
        m.team = request.form.get('team')
        m.identity = request.form.get('identity')
        m.status = request.form.get('status')
        session.commit()
        session.close()
        flash('修改成功', 'success')
        return redirect(url_for('admin_mechanics'))
    fields = [
        {"label": "姓名", "name": "name", "value": m.name or "", "type": "text"},
        {"label": "工号", "name": "employee_id", "value": m.employee_id or "", "type": "text"},
        {"label": "班组", "name": "team", "value": m.team or "", "type": "text"},
        {"label": "身份", "name": "identity", "value": m.identity or "", "type": "text"},
        {"label": "状态", "name": "status", "value": m.status or "", "type": "text"}
    ]
    session.close()
    return render_template('data_edit.html', title='编辑人员', fields=fields)

@app.route('/admin/mechanics/delete/<int:mid>')
@login_required
def delete_mechanic(mid):
    session = Session()
    m = session.query(Mechanic).get(mid)
    if m:
        session.query(Issue).filter(Issue.mechanic_id == mid).delete()
        session.query(MonthlyRecord).filter(MonthlyRecord.mechanic_id == mid).delete()
        session.query(RewardHistory).filter(RewardHistory.mechanic_id == mid).delete()
        session.query(ClearedHoursRecord).filter(ClearedHoursRecord.mechanic_id == mid).delete()
        session.delete(m)
        session.commit()
        flash('人员已删除', 'success')
    session.close()
    return redirect(url_for('admin_mechanics'))

@app.route('/issues/<int:mid>')
def view_issues(mid):
    session = Session()
    mechanic = session.query(Mechanic).get(mid)
    if getattr(current_user, 'is_authenticated', False) and current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        if mechanic and mechanic.workshop_id != current_user.workshop_id:
            session.close()
            flash('无权查看该人员问题记录', 'error')
            return redirect(url_for('mechanics'))
    from_source = request.args.get('from_source', '')
    issues = session.query(Issue).filter_by(mechanic_id=mid, status='未结算').all()
    if from_source == 'cleared_hours':
        ch = session.query(ClearedHoursRecord).filter_by(mechanic_id=mid).first()
        details = []
        if ch and ch.deduction_details:
            details = [s.strip() for s in str(ch.deduction_details).split('\n') if s and str(s).strip()]
        # 如果系统内没有扣分明细，尝试从“使用文件/附件8 … 清零汇总表.xlsx”的J列读取
        if not details:
            try:
                import openpyxl
                path = os.path.join('使用文件', '附件8 随车机械师积分及“千工时”保安全竞赛工时奖励清零汇总表.xlsx')
                if os.path.exists(path) and mechanic:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    ws = wb.active
                    # 查找表头行与列索引
                    def norm(s):
                        return (str(s).strip() if s is not None else '')
                    header_row = None
                    col_map = {}
                    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                        vals = [norm(v) for v in row]
                        if any('姓名' in v or '工号' in v for v in vals):
                            header_row = i
                            for idx, v in enumerate(vals):
                                if v:
                                    col_map[v] = idx
                            break
                    def find_idx(keys, default=None):
                        for k in col_map.keys():
                            if any(x in k for x in keys):
                                return col_map[k]
                        return default
                    idx_name = find_idx(['姓名'])
                    idx_emp = find_idx(['工号'])
                    # J列：扣分明细
                    idx_detail = find_idx(['扣分明细'])
                    if idx_detail is None:
                        idx_detail = 9  # 0-based -> 第10列(J)
                    if header_row:
                        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                            name_ok = (idx_name is not None and idx_name < len(row) and norm(row[idx_name]) == (mechanic.name or '').strip())
                            emp_ok = (idx_emp is not None and idx_emp < len(row) and norm(row[idx_emp]) == (mechanic.employee_id or '').strip())
                            if name_ok or emp_ok:
                                cell_val = row[idx_detail] if idx_detail < len(row) else None
                                if cell_val:
                                    s = str(cell_val).strip()
                                    details = [t.strip() for t in s.split('\n') if t.strip()]
                                break
            except Exception:
                pass
        if details:
            filtered = []
            for i in issues:
                p = (i.problem or '').strip()
                if any(d in p for d in details):
                    filtered.append(i)
            if filtered:
                issues = filtered
    session.close()
    return render_template('issues.html', mechanic=mechanic, issues=issues)

@app.route('/rewards')
@login_required
def rewards():
    session = Session()
    try:
        reward_month = request.args.get('reward_month')
        records_query = session.query(Attachment6Data)
        if current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
            records_query = records_query.join(Mechanic, Attachment6Data.employee_id == Mechanic.employee_id).filter(Mechanic.workshop_id == current_user.workshop_id)
        records = records_query.all()

        def normalize_month(val):
            if not val:
                return None
            s = str(val).strip()
            if not s:
                return None
            s = s.replace('年', '-').replace('月', '').replace('.', '-').replace('/', '-')
            parts = s.split('-')
            if len(parts) >= 2:
                y = parts[0]
                m = parts[1]
                if len(m) == 1:
                    m = f"0{m}"
                return f"{y}-{m}"
            if len(s) == 6 and s.isdigit():
                return f"{s[:4]}-{s[4:]}"
            return None

        def cycle_end_month(val):
            if not val:
                return None
            end_part = str(val).split('-')[-1].strip()
            return normalize_month(end_part)

        month_set = set()
        for r in records:
            m1 = normalize_month(r.reward_date)
            m2 = cycle_end_month(r.reward_cycle)
            if m1:
                month_set.add(m1)
            if m2:
                month_set.add(m2)
        reward_months = sorted(month_set, reverse=True)
        if not reward_month:
            reward_month = reward_months[0] if reward_months else datetime.now().strftime('%Y-%m')

        filtered = []
        for r in records:
            if normalize_month(r.reward_date) == reward_month or cycle_end_month(r.reward_cycle) == reward_month:
                filtered.append(r)
        filtered.sort(key=lambda r: (r.team or '', r.name or '', r.employee_id or ''))
        return render_template('rewards.html', records=filtered, reward_month=reward_month, reward_months=reward_months)
    except Exception as e:
        session.rollback()
        import traceback
        traceback.print_exc() # Print to console
        flash(f'加载奖励明细失败: {e}', 'error')
        return redirect(url_for('index'))
    finally:
        session.close()

@app.route('/export', methods=['POST'])
def export():
    session = Session()
    reward_month = request.form.get('reward_month', datetime.now().strftime('%Y-%m'))
    
    template_path = os.path.join('excel_templates', '附件5模板.xlsx')
    output_path = os.path.join('excel_templates', 'generated_rewards.xlsx')
    shutil.copy(template_path, output_path)
    
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Get font style from row 3 (which is likely correct)
    # Or just enforce standard font
    base_font = Font(name='宋体', size=11)
    
    row_idx = 3 # Start from row 3
    
    # Helper to calculate next cycle (Same as above)
    def get_next_cycle(m, target_month_str):
        try:
            # Target Date (End Month)
            target_date = datetime.strptime(target_month_str, '%Y-%m')
            target_fmt = f"{target_date.year}年{target_date.month}月"
            
            # Default Start
            start_fmt = "2025年12月" 
            
            if m.rewards_history:
                last_cycle = m.rewards_history[0].reward_cycle # e.g. 2025.07-2025.10
                if '-' in last_cycle:
                    # Extract end part of last cycle
                    parts_dash = last_cycle.split('-')
                    end_part = parts_dash[-1].strip() # Take the last part to be safe
                    
                    # New Logic: Start Month = End of Last Cycle + 1 Month
                    # Need to handle varied formats of end_part:
                    # "2025.10", "2025年10月", "2025-10"
                    
                    calc_dt = None
                    
                    # Clean up Chinese chars for parsing
                    clean_end = end_part.replace('年', '.').replace('月', '').strip()
                    
                    try:
                        # Try YYYY.MM
                        calc_dt = datetime.strptime(clean_end, '%Y.%m')
                    except:
                        try:
                            # Try YYYY-MM
                            calc_dt = datetime.strptime(clean_end, '%Y-%m')
                        except:
                            try:
                                # Manual split
                                if '.' in clean_end:
                                    ps = clean_end.split('.')
                                    calc_dt = datetime(int(ps[0]), int(ps[1]), 1)
                                elif '-' in clean_end:
                                    ps = clean_end.split('-')
                                    calc_dt = datetime(int(ps[0]), int(ps[1]), 1)
                            except:
                                pass
                    
                    if calc_dt:
                        # Add 1 month
                        if calc_dt.month == 12:
                            next_dt = datetime(calc_dt.year + 1, 1, 1)
                        else:
                            next_dt = datetime(calc_dt.year, calc_dt.month + 1, 1)
                        start_fmt = f"{next_dt.year}年{next_dt.month}月"
            
            return f"{start_fmt}-{target_fmt}"
        except Exception as e:
            print(f"Cycle Error: {e}")
            return f"ERROR-{target_month_str}"

    def normalize_month(val):
        if not val:
            return None
        s = str(val).strip()
        if not s:
            return None
        s = s.replace('年', '-').replace('月', '').replace('.', '-').replace('/', '-')
        parts = s.split('-')
        if len(parts) >= 2:
            y = parts[0]
            m = parts[1]
            if len(m) == 1:
                m = f"0{m}"
            return f"{y}-{m}"
        if len(s) == 6 and s.isdigit():
            return f"{s[:4]}-{s[4:]}"
        return None

    def cycle_end_month(val):
        if not val:
            return None
        end_part = str(val).split('-')[-1].strip()
        return normalize_month(end_part)

    def parse_number(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        for p in ['上期：', '上期:', '上期']:
            if s.startswith(p):
                s = s[len(p):].strip()
        try:
            return float(s)
        except:
            return None

    records_query = session.query(Attachment6Data)
    if getattr(current_user, 'is_authenticated', False) and current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
        records_query = records_query.join(Mechanic, Attachment6Data.employee_id == Mechanic.employee_id).filter(Mechanic.workshop_id == current_user.workshop_id)
    records_all = records_query.all()
    records = [r for r in records_all if normalize_month(r.reward_date) == reward_month or cycle_end_month(r.reward_cycle) == reward_month]

    if records:
        records_employee_ids = [r.employee_id for r in records if r.employee_id]
        employee_to_mid = {}
        if records_employee_ids:
            mech_rows = session.query(Mechanic.id, Mechanic.employee_id).filter(Mechanic.employee_id.in_(records_employee_ids)).all()
            employee_to_mid = {r[1]: r[0] for r in mech_rows}
        mids = list({mid for mid in employee_to_mid.values() if mid})
        prev_extra_map = {}
        if mids:
            hist_rows = session.query(RewardHistory.mechanic_id, RewardHistory.reward_date, RewardHistory.extra_reward).filter(
                RewardHistory.mechanic_id.in_(mids)
            ).order_by(RewardHistory.mechanic_id.asc(), RewardHistory.reward_date.asc()).all()
            per_mid = {}
            for mid, rdate, extra in hist_rows:
                mkey = normalize_month(rdate) or str(rdate or '').strip()
                if not mkey:
                    continue
                per_mid.setdefault(mid, []).append((mkey, float(extra or 0.0)))
            for mid, arr in per_mid.items():
                seen = {}
                for mkey, extra in arr:
                    seen[mkey] = extra
                keys_sorted = sorted(seen.keys())
                prev = 0.0
                for k in keys_sorted:
                    prev_extra_map[(mid, k)] = prev
                    prev = float(seen[k] or 0.0)

        records.sort(key=lambda r: (r.team or '', r.name or '', r.employee_id or ''))
        for i, r in enumerate(records, 1):
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=r.name)
            ws.cell(row=row_idx, column=3, value=r.employee_id)
            ws.cell(row=row_idx, column=4, value=r.team)
            ws.cell(row=row_idx, column=5, value=r.activity_cumulative_hours or 0.0)
            ws.cell(row=row_idx, column=6, value=r.reward_date or reward_month)
            ws.cell(row=row_idx, column=7, value=r.reward_cycle)
            ws.cell(row=row_idx, column=8, value=r.cycle_deduction or 0.0)
            ws.cell(row=row_idx, column=9, value=r.reward_amount or 0.0)
            ws.cell(row=row_idx, column=10, value=r.cleared_hours or 0.0)
            ws.cell(row=row_idx, column=11, value=r.balance_hours or 0.0)
            ws.cell(row=row_idx, column=12, value=r.past_reward_count or 0)
            ws.cell(row=row_idx, column=13, value=r.is_consecutive or '')
            mid = employee_to_mid.get(r.employee_id)
            mkey = normalize_month(r.reward_date) or cycle_end_month(r.reward_cycle) or str(r.reward_date or '').strip()
            prev_extra = prev_extra_map.get((mid, mkey), parse_number(r.consecutive_info) or 0.0)
            ws.cell(row=row_idx, column=14, value=prev_extra)
            ws.cell(row=row_idx, column=15, value=r.extra_reward or 0.0)
            ws.cell(row=row_idx, column=16, value=r.total_amount or 0.0)
            for col in range(1, 17):
                cell = ws.cell(row=row_idx, column=col)
                current_font = cell.font
                if current_font:
                    new_font = current_font.copy(name='宋体', size=11)
                    cell.font = new_font
                else:
                    cell.font = Font(name='宋体', size=11)
            row_idx += 1
    else:
        mechanics_q = session.query(Mechanic).filter(
            Mechanic.total_hours >= 1000,
            Mechanic.identity == '随车机械师',
            Mechanic.status == '在岗'
        )
        if getattr(current_user, 'is_authenticated', False) and current_user.role == ROLE_WORKSHOP and current_user.workshop_id:
            mechanics_q = mechanics_q.filter(Mechanic.workshop_id == current_user.workshop_id)
        mechanics = mechanics_q.order_by(Mechanic.name.asc()).options(joinedload(Mechanic.rewards_history)).all()

        for i, m in enumerate(mechanics, 1):
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=m.name)
            ws.cell(row=row_idx, column=3, value=m.employee_id)
            ws.cell(row=row_idx, column=4, value=m.team)
            
            base = m.base_hours if m.base_hours is not None else 0.0
            act_cumulative = (len(m.rewards_history) * 1000) + base
            ws.cell(row=row_idx, column=5, value=act_cumulative)
            
            try:
                dt = datetime.strptime(reward_month, '%Y-%m')
                ws.cell(row=row_idx, column=6, value=dt.strftime('%Y.%m'))
            except:
                ws.cell(row=row_idx, column=6, value=reward_month)
            
            cycle_str = get_next_cycle(m, reward_month)
            ws.cell(row=row_idx, column=7, value=cycle_str)
            
            cur_ded = m.current_cycle_deduction if m.current_cycle_deduction is not None else 0.0
            ws.cell(row=row_idx, column=8, value=-cur_ded if cur_ded > 0 else 0)
            
            reward_amt = 0
            if cur_ded == 0:
                reward_amt = 1200
            elif cur_ded <= 1:
                reward_amt = 1000
            elif cur_ded < 3:
                reward_amt = 600
            elif cur_ded < 6:
                reward_amt = 200
            else:
                reward_amt = 0
                
            ws.cell(row=row_idx, column=9, value=reward_amt)
            
            cleared_rec = session.query(ClearedHoursRecord).filter_by(mechanic_id=m.id).first()
            cleared_val = 0.0
            if cleared_rec:
                cleared_val = cleared_rec.total_cleared_hours
                
            ws.cell(row=row_idx, column=10, value=cleared_val)
            
            total_h = m.total_hours if m.total_hours is not None else 0.0
            ws.cell(row=row_idx, column=11, value=total_h - 1000)
            
            last_reward = m.rewards_history[0] if m.rewards_history else None
            last_amt = last_reward.amount if last_reward else 0
            ws.cell(row=row_idx, column=14, value=last_amt)
            
            past_count = len(m.rewards_history) if m.rewards_history else 0
            ws.cell(row=row_idx, column=12, value=past_count)
            
            is_consecutive = "否"
            extra_amt = 0
            cur_ded_val = -abs(m.current_cycle_deduction or 0.0)
            if last_reward:
                last_ded_val = -abs(last_reward.deduction or 0.0)
                last_triggered = (last_reward.extra_reward or 0) > 0
                if cur_ded_val <= -1 or last_ded_val <= -1 or last_triggered:
                    is_consecutive = "否"
                    extra_amt = 0
                else:
                    is_consecutive = "是"
                    if cur_ded_val == 0 and last_ded_val == 0:
                        extra_amt = 2000
                    else:
                        extra_amt = 1500
            else:
                is_consecutive = "否"
                extra_amt = 0
            
            ws.cell(row=row_idx, column=13, value=is_consecutive)
            ws.cell(row=row_idx, column=15, value=extra_amt)
            ws.cell(row=row_idx, column=16, value=reward_amt + extra_amt)
            
            for col in range(1, 17):
                cell = ws.cell(row=row_idx, column=col)
                current_font = cell.font
                if current_font:
                    new_font = current_font.copy(name='宋体', size=11)
                    cell.font = new_font
                else:
                    cell.font = Font(name='宋体', size=11)
            
            row_idx += 1
        
    wb.save(output_path)
    session.close()
    return send_file(output_path, as_attachment=True, download_name="奖励明细.xlsx")


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
